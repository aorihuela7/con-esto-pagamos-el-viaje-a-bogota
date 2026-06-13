import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sg=json.load(open("segura.json"))
SG_N=len(sg["combos"]); SG_COST=int(sg["cost"])
SG_ALL=sum(c["profit"] for c in sg["combos"])      # ganancia si TODAS ganan
SG_EV=round(sg["agg_ev"])
AR="Arial"; NAVY="1F3864"; BLUE="2E5496"; TEAL="1F6E6E"; GOLD="BF9000"; GREEN="1F6E43"
LBLUE="DDEBF7"; LGOLD="FFF2CC"; LGREEN="C6EFCE"; MGREEN="E2EFDA"; LGREY="F2F2F2"
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
rgt=Alignment(horizontal="right",vertical="center",wrap_text=True)
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=load_workbook("Sistema_Mundial_2026_COMPLETO.xlsx")
if "Portada" in wb.sheetnames: del wb["Portada"]
p=wb.create_sheet("Portada"); wb.move_sheet("Portada", -(len(wb.sheetnames)-1))
p.sheet_view.showGridLines=False
for col,w in zip("ABCDEFG",[2,26,13,19,21,34,2]): p.column_dimensions[col].width=w

p.merge_cells("B2:F2")
c=p["B2"]; c.value="SISTEMA DE APUESTAS · MUNDIAL 2026"; c.font=f(16,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; p.row_dimensions[2].height=32
p.merge_cells("B3:F3")
c=p["B3"]; c.value="Cuatro estrategias en un archivo. Abajo: cuánto inviertes, cuánto puedes ganar y dónde conviene poner el dinero."; c.font=f(10,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; p.row_dimensions[3].height=22

# ---------- TABLA INVERSION / GANANCIA ----------
r=5
p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
c=p.cell(r,2,"INVERSIÓN Y GANANCIA POR ESTRATEGIA"); c.font=f(12,True,"FFFFFF"); c.fill=fill(GREEN); c.alignment=lft; p.row_dimensions[r].height=24; r+=1
heads=["Estrategia","Inversión","Si TODO sale bien","Resultado esperado","Riesgo / nota"]
for k,h in enumerate(heads):
    cc=p.cell(r,2+k,h); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(NAVY); cc.alignment=ctr; cc.border=bd
p.row_dimensions[r].height=24; r+=1
rows=[
 ("1 · Combinadas iniciales","S/ 240","+S/ 46,283","+S/ 683 a +S/ 2,848 / ronda (según tu confianza)",
  "ALTO — cada combinada necesita 20 aciertos; a cuotas de mercado queda en break-even.",LBLUE),
 ("2 · Kelly (por día)","Banca a tu elección (ej. S/ 100)","La banca crece cada día",
  "Mediana ×2.2 (½), ×2.6 (⅔), ×3.1 (100%) en 16 días; ganas 79–88% de los casos",
  "BAJO-MEDIO — único método diseñado para NO quebrar.",LGOLD),
 ("3 · Combinadas +S/ 5,000","S/ 540  (18 × S/ 30)","+S/ 5,000 por cada acierto",
  "+S/ 4,366 promedio, pero solo 35% de prob. de ganar algo",
  "MUY ALTO — lo más probable: perder casi todo. Billete de lotería con ventaja.",LGREEN),
 (f"4 · Combinada Segura (Betsson)",f"S/ {SG_COST:,}  ({SG_N} × S/ 60)",f"+S/ {SG_ALL:,}  (las {SG_N} ganan)",
  f"+S/ {SG_EV:,} (EV con el seguro). Si falla 1 pata por combinada, te la reembolsan",
  "MEDIO — el seguro convierte un fallo en empate.",MGREEN),
]
for est,inv,bien,esp,riesgo,col in rows:
    cc=p.cell(r,2,est); cc.font=f(9,True,NAVY); cc.alignment=lft; cc.border=bd; cc.fill=fill(col)
    cc=p.cell(r,3,inv); cc.font=f(9,True); cc.alignment=ctr; cc.border=bd
    cc=p.cell(r,4,bien); cc.font=f(9,True,"006100"); cc.alignment=ctr; cc.border=bd
    cc=p.cell(r,5,esp); cc.font=f(9); cc.alignment=lft; cc.border=bd
    cc=p.cell(r,6,riesgo); cc.font=f(9); cc.alignment=lft; cc.border=bd
    p.row_dimensions[r].height=46; r+=1
# total invertible
p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
cc=p.cell(r,2,"Inversión total (sin contar Kelly)"); cc.font=f(9,True); cc.alignment=rgt; cc.border=bd; cc.fill=fill(LGREY)
cc=p.cell(r,3); cc.border=bd; cc.fill=fill(LGREY)
cc=p.cell(r,4,f"S/ {240+540+SG_COST:,}"); cc.font=f(10,True,NAVY); cc.alignment=ctr; cc.border=bd; cc.fill=fill(LGREY)
p.merge_cells(start_row=r,start_column=5,end_row=r,end_column=6)
cc=p.cell(r,5,f"240 + 540 + {SG_COST}. Kelly aparte porque usa una banca, no un ticket fijo."); cc.font=f(8,False,"808080"); cc.alignment=lft; cc.border=bd; cc.fill=fill(LGREY)
p.row_dimensions[r].height=20; r+=2

# ---------- RECOMENDACIONES ----------
p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
c=p.cell(r,2,"RECOMENDACIÓN — ¿dónde poner el dinero?"); c.font=f(12,True,"FFFFFF"); c.fill=fill(GOLD); c.alignment=lft; p.row_dimensions[r].height=22; r+=1
recs=[
 ("Crecer sin quebrar → KELLY (½ o ⅔).","Es la única estrategia pensada para sobrevivir las rachas malas: apuesta un % de la banca, nunca todo. Lo más sólido si quieres que el dinero dure y crezca."),
 ("Aprovechar la promo con riesgo acotado → COMBINADA SEGURA.","El reembolso de Betsson convierte el fallo de 1 pata en empate, así que el riesgo real es menor que el que sugiere su pago. Buen segundo lugar."),
 ("Solo diversión → COMBINADAS DE +S/ 5,000.","Billetes de lotería con ventaja: mete poco (asume que esos S/540 los puedes perder). El premio gordo paga, pero lo más probable es no acertar."),
 ("Combinadas iniciales → solo si confías de verdad en tus probabilidades.","A cuotas de mercado quedan en break-even; su valor depende por completo de que tu lectura sea mejor que la de la casa."),
 ("Reparto ilustrativo (si tuvieras S/ 1,000):","~S/ 400 a Kelly, ~S/ 500 a la Combinada Segura, ~S/ 100 a una de las +S/ 5,000. Ajusta a tu tolerancia al riesgo."),
]
for a,b in recs:
    p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
    ca=p.cell(r,2,a); ca.font=f(9,True,NAVY); ca.alignment=lft; ca.border=bd; ca.fill=fill(LGOLD)
    p.merge_cells(start_row=r,start_column=4,end_row=r,end_column=6)
    cb=p.cell(r,4,b); cb.font=f(9); cb.alignment=lft; cb.border=bd
    p.row_dimensions[r].height=40; r+=1
p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
c=p.cell(r,2,"No soy asesor financiero: esto es información para que decidas tú. Apuesta solo lo que puedas permitirte perder.")
c.font=f(9,True,"808080"); c.alignment=lft; p.row_dimensions[r].height=18; r+=2

# ---------- INDICE DE HOJAS ----------
p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
c=p.cell(r,2,"ÍNDICE DE HOJAS"); c.font=f(12,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=lft; p.row_dimensions[r].height=22; r+=1
blocks=[
 ("1 · LAS COMBINADAS INICIALES", TEAL, [
   ("Inicio · Resumen","Visión general del primer sistema de combinadas."),
   ("Inicio · Combinadas","Las combinadas armadas al principio, con prob. de la cuota y la tuya."),
   ("Inicio · Cobertura","Qué partidos cubre cada combinada."),
 ]),
 ("2 · KELLY", GOLD, [
   ("Kelly · Por día","Combinada por día con la sugerencia de Kelly ½, ⅔ y 100%."),
   ("Kelly · Comparación","½ vs ⅔ vs 100%: mediana, prob. de ganar, de quebrar, P10, P90."),
   ("Kelly · Percentiles","Distribución P10–P90 de la banca final para los 3 Kelly."),
   ("Kelly ½/⅔/100% · gana todo","Trayectoria de la banca si ganas siempre con cada fracción."),
 ]),
 ("3 · COMBINADAS DE +S/ 5,000", BLUE, [
   ("5000 · Resumen y riesgo","18 combinadas de S/30 que pagan +S/5,000, con su análisis de riesgo."),
   ("5000 · Detalle","Los partidos que forman cada una de las 18 combinadas."),
   ("5000 · Mapa","Mapa visual de diversificación."),
 ]),
 ("4 · COMBINADA SEGURA (BETSSON)", GREEN, [
   ("Segura · Resumen","12 combinadas de S/60 con el reembolso de Betsson + diversificación."),
   ("Segura · Detalle","Los 5 partidos de cada una de las 12 combinadas seguras."),
   ("Segura · Mapa","Mapa de diversificación + ganancia por combinada apostando S/60."),
 ]),
]
for title,col,items in blocks:
    p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    c=p.cell(r,2,title); c.font=f(11,True,"FFFFFF"); c.fill=fill(col); c.alignment=lft; p.row_dimensions[r].height=22; r+=1
    for name,desc in items:
        cl=p.cell(r,2,name); cl.font=f(9,True,NAVY); cl.alignment=lft; cl.border=bd; cl.fill=fill(LBLUE)
        p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=2)
        cd=p.cell(r,3,desc); cd.font=f(9); cd.alignment=lft; cd.border=bd
        p.merge_cells(start_row=r,start_column=3,end_row=r,end_column=6)
        p.row_dimensions[r].height=20; r+=1
    r+=1

wb.save("Sistema_Mundial_2026_COMPLETO.xlsx")
print("portada reconstruida; orden hojas:")
print(wb.sheetnames)

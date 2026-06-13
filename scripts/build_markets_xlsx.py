from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AR="Arial"; NAVY="1F3864"; BLUE="2E5496"; TEAL="1F6E6E"; GOLD="BF9000"
LGREY="F2F2F2"; LBLUE="DDEBF7"; LGOLD="FFF2CC"; YEL="FFFDE7"; GREEN="C6EFCE"; RED="F4CCCC"
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=Workbook(); s=wb.active; s.title="Mercados Alt."; s.sheet_view.showGridLines=False
# cols: A sp, B Partido, C Mercado, D Prom Eq1, E Prom Eq2, F Línea, G λ, H P(Over), I Cuota justa O, J Cuota casa, K Valor %, L ¿Apostar?
widths=[2,24,16,9,9,7,8,9,11,10,9,12]
for i,w in enumerate(widths): s.column_dimensions[get_column_letter(i+1)].width=w

s.merge_cells("B2:L2")
c=s["B2"]; c.value="CALCULADORA DE MERCADOS ALTERNATIVOS  (modelo Poisson)"; c.font=f(15,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s.row_dimensions[2].height=28
s.merge_cells("B3:L3")
c=s["B3"]; c.value=("Goles · Córners · Tarjetas · Tiros.  Pones el promedio de cada equipo (de Sofascore) y el modelo calcula la probabilidad real de superar la línea, "
 "la cuota justa y si la cuota de la casa tiene VALOR.")
c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=30

# header
r=5
hdr=["Partido","Mercado","Prom Eq1","Prom Eq2","Línea","λ esperado","P(Over)","Cuota justa","Cuota casa","Valor %","¿Apostar?"]
for j,h in enumerate(hdr):
    cc=s.cell(r,2+j,h); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(BLUE); cc.alignment=ctr; cc.border=bd
s.row_dimensions[r].height=30
r+=1
first=r

# example rows: (partido, mercado, prom1, prom2, linea, cuota_casa)  -- placeholders to replace
rows=[
 ("Brasil vs Haití","Goles O","2.30","0.50","2.5","1.55"),
 ("España vs Cabo Verde","Goles O","2.40","0.60","2.5","1.50"),
 ("México vs Sudáfrica","Goles O","1.60","0.90","2.5","1.95"),
 ("Argentina vs Argelia","Córners O","6.00","3.50","8.5","1.85"),
 ("Inglaterra vs Ghana","Córners O","6.50","3.00","8.5","1.80"),
 ("Croacia vs Panamá","Córners O","5.50","3.50","8.5","1.90"),
 ("Marruecos vs Haití","Tarjetas O","2.20","2.40","3.5","1.80"),
 ("Argentina vs Austria","Tarjetas O","2.30","2.10","4.5","2.05"),
 ("Bélgica vs Irán","Tarjetas O","2.00","2.60","3.5","1.75"),
 ("Brasil vs Haití","Tiros puerta O","6.00","2.50","7.5","1.85"),
 ("Francia vs Senegal","Tiros puerta O","5.50","3.00","7.5","1.90"),
 ("Alemania vs Curazao","Tiros puerta O","7.00","2.00","8.5","1.95"),
]
def setrow(rr, vals, example=True):
    part,merc,p1,p2,line,cuota=vals
    s.cell(rr,2,part); s.cell(rr,3,merc)
    s.cell(rr,4,float(p1)); s.cell(rr,5,float(p2)); s.cell(rr,6,float(line)); s.cell(rr,10,float(cuota))
    s.cell(rr,7,f"=D{rr}+E{rr}").number_format="0.00"                       # lambda
    s.cell(rr,8,f'=1-SUMPRODUCT(EXP(-G{rr})*G{rr}^(ROW(INDIRECT("1:"&(ROUNDDOWN(F{rr},0)+1)))-1)/FACT(ROW(INDIRECT("1:"&(ROUNDDOWN(F{rr},0)+1)))-1))').number_format="0.0%"  # P(Over)
    s.cell(rr,9,f'=IF(H{rr}>0,1/H{rr},"")').number_format="0.00"            # cuota justa
    s.cell(rr,11,f"=J{rr}*H{rr}-1").number_format="0.0%"                    # valor
    s.cell(rr,12,f'=IF(J{rr}="","",IF(K{rr}>0.03,"VALOR ✓",IF(K{rr}>=0,"justo","evitar")))')
    for col in range(2,13):
        cell=s.cell(rr,col); cell.border=bd; cell.font=f(10); cell.alignment=lft if col in(2,3) else ctr
        cell.fill=fill("FFFFFF")
    for col in (4,5,6,10):  # inputs
        s.cell(rr,col).fill=fill(YEL); s.cell(rr,col).font=f(10,False,"0000FF")
    s.cell(rr,4).number_format="0.00"; s.cell(rr,5).number_format="0.00"
    s.cell(rr,6).number_format="0.0"; s.cell(rr,10).number_format="0.00"

for i,v in enumerate(rows):
    setrow(first+i, v)
# blank rows to add own
nblank=8
for k in range(nblank):
    rr=first+len(rows)+k
    s.cell(rr,7,f"=IF(AND(D{rr}<>\"\",E{rr}<>\"\"),D{rr}+E{rr},\"\")").number_format="0.00"
    s.cell(rr,8,f'=IF(OR(F{rr}="",G{rr}=""),"",1-SUMPRODUCT(EXP(-G{rr})*G{rr}^(ROW(INDIRECT("1:"&(ROUNDDOWN(F{rr},0)+1)))-1)/FACT(ROW(INDIRECT("1:"&(ROUNDDOWN(F{rr},0)+1)))-1)))').number_format="0.0%"
    s.cell(rr,9,f'=IF(H{rr}="","",1/H{rr})').number_format="0.00"
    s.cell(rr,11,f'=IF(OR(H{rr}="",J{rr}=""),"",J{rr}*H{rr}-1)').number_format="0.0%"
    s.cell(rr,12,f'=IF(OR(K{rr}="",J{rr}=""),"",IF(K{rr}>0.03,"VALOR ✓",IF(K{rr}>=0,"justo","evitar")))')
    for col in range(2,13):
        cell=s.cell(rr,col); cell.border=bd; cell.font=f(10); cell.alignment=lft if col in(2,3) else ctr; cell.fill=fill("FFFFFF")
    for col in (4,5,6,10):
        s.cell(rr,col).fill=fill(YEL); s.cell(rr,col).font=f(10,False,"0000FF")
    s.cell(rr,4).number_format="0.00"; s.cell(rr,5).number_format="0.00"; s.cell(rr,6).number_format="0.0"; s.cell(rr,10).number_format="0.00"
last=first+len(rows)+nblank-1

# conditional formatting for valor column via simple fill on examples already computed at recalc? leave text flag.
r=last+2
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=12)
c=s.cell(r,2,"Cómo usarla"); c.font=f(12,True,NAVY); c.fill=fill(LGOLD); c.alignment=lft; r+=1
notes=[
 "Las filas de ejemplo traen promedios inventados (azul). Reemplázalos con los reales de Sofascore: pestaña del equipo → promedios por partido.",
 "Prom Eq1 / Prom Eq2: para GOLES, lo que anota cada equipo. Para CÓRNERS/TARJETAS/TIROS, el promedio de cada equipo en ese evento. El modelo suma ambos (λ).",
 "Línea: la del mercado (2.5 goles, 8.5 córners, 3.5 tarjetas…). Pon la cuota de tu casa en 'Cuota casa'.",
 "P(Over) = probabilidad real de superar la línea (Poisson). Cuota justa = 1 ÷ P(Over). Si la cuota de la casa es MAYOR que la justa → hay VALOR ✓.",
 "Valor % = cuánto ganas en promedio por sol apostado. Solo apuesta las filas con 'VALOR ✓'. Combínalas con Kelly de las otras hojas para el monto.",
 "Tarjetas dependen mucho del árbitro: ajusta el promedio al alza en clásicos o árbitros estrictos. Córners y tiros son los más estables.",
]
for t in notes:
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=12)
    c=s.cell(r,2,"•  "+t); c.font=f(10); c.alignment=lft; s.row_dimensions[r].height=28; r+=1
r+=1
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=12)
c=s.cell(r,2,"Por qué estos mercados son 'más seguros': no dependen de quién gana, sino de promedios que se repiten partido a partido. Por eso la varianza es menor que en el 1X2; pero ojo, cuota más baja = premio más chico. El valor manda, no la 'seguridad' aparente.")
c.font=f(9,False,"808080"); c.alignment=lft; s.row_dimensions[r].height=44

s.freeze_panes="B6"
wb.save("Mercados_Alternativos.xlsx")
print("saved rows",len(rows),"+",nblank,"blank")

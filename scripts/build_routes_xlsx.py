import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d = json.load(open("routes.json"))
routes = d["routes"]; agg = d["agg"]

AR = "Arial"
NAVY = "1F3864"; BLUE = "2E5496"; GOLD = "BF9000"; GREEN = "C6EFCE"
LGREY = "F2F2F2"; LBLUE = "DDEBF7"; LGOLD = "FFF2CC"; RED = "FFC7CE"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
def f(sz=10, b=False, color="000000"): return Font(name=AR, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ============ RESUMEN ============
s = wb.active; s.title = "Resumen"
s.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGH", [3, 22, 16, 16, 16, 16, 16, 3]):
    s.column_dimensions[col].width = w

s.merge_cells("B2:G2")
c = s["B2"]; c.value = "RUTAS — Mundial 2026"; c.font = f(18, True, "FFFFFF"); c.fill = fill(NAVY); c.alignment = center
s.row_dimensions[2].height = 30
s.merge_cells("B3:G3")
c = s["B3"]; c.value = "6 escaleras de apuestas encadenadas · banca total S/120 · cobra (cashout) cuando la ruta llegue a su pata recomendada"
c.font = f(10, False, "FFFFFF"); c.fill = fill(BLUE); c.alignment = center
s.row_dimensions[3].height = 26

r = 5
s.merge_cells(f"B{r}:G{r}")
c = s[f"B{r}"]; c.value = "¿Cómo funciona?"; c.font = f(12, True, NAVY); c.fill = fill(LBLUE); c.alignment = left
r += 1
how = [
 "Cada RUTA es una combinada que vas armando partido a partido. Empiezas apostando S/20 en la 1ra pata.",
 "Si gana, la ganancia rueda automáticamente a la siguiente pata (no agregas dinero nuevo).",
 "La columna 'Si cierras aquí' te dice cuánto cobras si haces cashout en esa pata.",
 "La zona dorada = pata recomendada para cobrar (donde la probabilidad de seguir vivo aún es ≈50% o más).",
 "Marca 1 (ganó) o 0 (perdió) en la columna ¿Ganó? a medida que se juegan. El Estado se actualiza solo.",
 "Si una ruta muere temprano, las otras 5 siguen vivas — por eso se diversifica en 6 rutas paralelas.",
 "Total arriesgado fijo: S/120 (S/20 × 6). No puedes perder más que eso.",
]
for t in how:
    s.merge_cells(f"B{r}:G{r}")
    cc = s[f"B{r}"]; cc.value = "•  " + t; cc.font = f(10); cc.alignment = left
    s.row_dimensions[r].height = 16; r += 1

r += 1
s.merge_cells(f"B{r}:G{r}")
c = s[f"B{r}"]; c.value = "Expectativa del portafolio (200k simulaciones, cobrando en la pata recomendada)"
c.font = f(12, True, NAVY); c.fill = fill(LBLUE); c.alignment = left
r += 1
hdr = ["Métrica", "Con TUS probabilidades", "Con prob. de mercado"]
for j, h in enumerate(hdr):
    cc = s.cell(r, 2 + j, h); cc.font = f(10, True, "FFFFFF"); cc.fill = fill(BLUE); cc.alignment = center; cc.border = border
r += 1
mine, mkt = agg["mine"], agg["market"]
rows = [
 ("Arriesgado total", f"S/ {mine['staked']:.0f}", f"S/ {mkt['staked']:.0f}"),
 ("Retorno promedio", f"S/ {mine['avg_ret']:.0f}", f"S/ {mkt['avg_ret']:.0f}"),
 ("Neto promedio", f"S/ {mine['avg_net']:+.0f}", f"S/ {mkt['avg_net']:+.0f}"),
 ("P(alguna ruta cobra)", f"{mine['p_any']:.0f} %", f"{mkt['p_any']:.0f} %"),
 ("P(ronda rentable)", f"{mine['p_profit']:.0f} %", f"{mkt['p_profit']:.0f} %"),
 ("Escenario malo (P10)", f"S/ {mine['p10']:+.0f}", f"S/ {mkt['p10']:+.0f}"),
 ("Escenario bueno (P90)", f"S/ {mine['p90']:+.0f}", f"S/ {mkt['p90']:+.0f}"),
]
for i, (a, b, cc_) in enumerate(rows):
    bg = LGREY if i % 2 else "FFFFFF"
    for j, v in enumerate((a, b, cc_)):
        cell = s.cell(r, 2 + j, v); cell.font = f(10, j == 0); cell.fill = fill(bg)
        cell.alignment = left if j == 0 else center; cell.border = border
    r += 1

r += 1
s.merge_cells(f"B{r}:G{r}")
c = s[f"B{r}"]; c.value = "Nota: con probabilidades de mercado el neto es ≈0 (es lo justo). El borde positivo depende de que tus probabilidades sean reales. Trátalo como apuesta de entretenimiento, no inversión."
c.font = f(9, False, "808080"); c.alignment = left; s.row_dimensions[r].height = 28

# ============ RUTAS ============
s2 = wb.create_sheet("Rutas")
s2.sheet_view.showGridLines = False
widths = [3, 6, 9, 26, 8, 9, 10, 9, 14, 10, 16]
for i, w in enumerate(widths):
    s2.column_dimensions[get_column_letter(i + 1)].width = w
# cols: A spacer, B Pata, C Fecha, D Pick, E Cuota, F Mi P, G Mult, H Vivo%, I Si cierras S/, J ¿Ganó?, K Estado

row = 2
COLHDR = ["Pata", "Fecha", "Pick", "Cuota", "Mi P", "Mult", "Vivo %", "Si cierras S/", "¿Ganó? (1/0)", "Estado"]
for rt in routes:
    n = len(rt["legs"]); s0 = rt["stake"]; rec = rt["rec_exit_idx"]
    # title
    s2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
    c = s2.cell(row, 2, f"RUTA {rt['route']}   ·   apuesta inicial S/ {s0:.0f}   ·   pata recomendada para cobrar: #{rec+1}")
    c.font = f(12, True, "FFFFFF"); c.fill = fill(NAVY); c.alignment = left
    s2.row_dimensions[row].height = 22
    row += 1
    # header
    for j, h in enumerate(COLHDR):
        cc = s2.cell(row, 2 + j, h); cc.font = f(9, True, "FFFFFF"); cc.fill = fill(BLUE); cc.alignment = center; cc.border = border
    s2.row_dimensions[row].height = 26
    hdr_row = row
    row += 1
    first_data = row
    res_col = "J"  # ¿Ganó?
    for k, lg in enumerate(rt["legs"]):
        rr = row
        res_range = f"${res_col}${first_data}:${res_col}{rr}"
        estado = (f'=IF(COUNTIF({res_range},0)>0,"MUERTA",'
                  f'IF({res_col}{rr}=1,"VIVA",IF(COUNT({res_range})={k+1},"—","—")))')
        vals = [k + 1, lg["date"], lg["pick"], lg["odds"], lg["prob"],
                lg["mult"], lg["surv"], lg["stop_payout"], None, estado]
        is_rec = (k == rec)
        for j, v in enumerate(vals):
            cc = s2.cell(rr, 2 + j, v); cc.border = border
            cc.font = f(10, is_rec)
            if j in (0, 1, 3, 4, 5, 6, 7, 8, 9):
                cc.alignment = center
            else:
                cc.alignment = left
            # number formats
            colidx = j  # 0=pata
            if j == 3: cc.number_format = "0.00"       # cuota
            if j == 4: cc.number_format = "0%"          # mi p
            if j == 5: cc.number_format = "0.00x"       # mult  -> use 0.00"x"
            if j == 6: cc.number_format = "0%"          # vivo
            if j == 7: cc.number_format = '"S/ "0.00'   # si cierras
            bg = LGOLD if is_rec else (LGREY if k % 2 else "FFFFFF")
            cc.fill = fill(bg)
        # fix mult format (openpyxl needs literal)
        s2.cell(rr, 7).number_format = '0.00"x"'
        # ¿Ganó? input cell -> light highlight
        s2.cell(rr, 11).fill = fill("FFFDE7")
        row += 1
    # route summary line
    res_range_full = f"${res_col}${first_data}:${res_col}${row-1}"
    rec_payrow = first_data + rec
    s2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    cc = s2.cell(row, 2,
        f'=IF(COUNTIF({res_range_full},0)>0,"✗ Ruta muerta",'
        f'IF(COUNT({res_range_full})=0,"Sin jugar aún",'
        f'"En curso — patas ganadas: "&COUNTIF({res_range_full},1)))')
    cc.font = f(10, True, NAVY); cc.fill = fill(LBLUE); cc.alignment = left; cc.border = border
    # cobro recomendado
    cc2 = s2.cell(row, 9, f"=I{rec_payrow}"); cc2.number_format = '"S/ "0.00'
    cc2.font = f(10, True, GOLD); cc2.fill = fill(LGOLD); cc2.alignment = center; cc2.border = border
    s2.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
    cc3 = s2.cell(row, 10, "← cobro en pata recom."); cc3.font = f(9, False, "808080"); cc3.alignment = left
    row += 2

# ============ CASHOUT Y HEDGE ============
s3 = wb.create_sheet("Cashout y Hedge")
s3.sheet_view.showGridLines = False
for col, w in zip("ABCDEF", [3, 30, 16, 16, 16, 3]):
    s3.column_dimensions[col].width = w
s3.merge_cells("B2:E2")
c = s3["B2"]; c.value = "Calculadora de Cobertura (Hedge)"; c.font = f(16, True, "FFFFFF"); c.fill = fill(NAVY); c.alignment = center
s3.row_dimensions[2].height = 28
s3.merge_cells("B3:E3")
c = s3["B3"]; c.value = "¿Tu ruta está viva y queda 1 pata? Apuesta al rival para asegurar ganancia pase lo que pase."
c.font = f(10, False, "FFFFFF"); c.fill = fill(BLUE); c.alignment = center
s3.row_dimensions[3].height = 22

inp = [
 ("Ganancia potencial si ganas (W = mult acum × S/20)", 83.26, '"S/ "0.00', LGOLD),
 ("Cuota del resultado CONTRARIO (la última pata)", 3.40, "0.00", LGOLD),
]
r = 5
s3.merge_cells(f"B{r}:E{r}"); c = s3[f"B{r}"]; c.value = "Ingresa tus datos (celdas amarillas)"; c.font = f(11, True, NAVY); c.fill = fill(LBLUE); c.alignment = left
r += 1
for label, val, fmt, bg in inp:
    s3.cell(r, 2, label).font = f(10); s3.cell(r, 2).alignment = left; s3.cell(r, 2).border = border
    cc = s3.cell(r, 3, val); cc.font = f(10, True, "0000FF"); cc.fill = fill(bg); cc.number_format = fmt; cc.alignment = center; cc.border = border
    r += 1
W = "C6"; OC = "C7"
r += 1
s3.merge_cells(f"B{r}:E{r}"); c = s3[f"B{r}"]; c.value = "Resultado"; c.font = f(11, True, NAVY); c.fill = fill(LBLUE); c.alignment = left
r += 1
outs = [
 ("Monto a apostar al contrario (hedge = W / cuota)", f"={W}/{OC}", '"S/ "0.00'),
 ("Ganancia asegurada si gana la ruta", f"={W}-({W}/{OC})", '"S/ "0.00'),
 ("Ganancia asegurada si gana el contrario", f"=({W}/{OC})*{OC}-({W}/{OC})", '"S/ "0.00'),
]
for label, formula, fmt in outs:
    s3.cell(r, 2, label).font = f(10); s3.cell(r, 2).alignment = left; s3.cell(r, 2).border = border
    cc = s3.cell(r, 3, formula); cc.font = f(10, True); cc.number_format = fmt; cc.alignment = center; cc.border = border; cc.fill = fill(GREEN)
    r += 1
r += 1
s3.merge_cells(f"B{r}:E{r}")
c = s3[f"B{r}"]; c.value = "Si las dos 'ganancias aseguradas' son positivas, tienes profit garantizado sin importar el resultado. Compara siempre contra el cashout que ofrece la casa."
c.font = f(9, False, "808080"); c.alignment = left; s3.row_dimensions[r].height = 30

wb.save("Rutas_Mundial.xlsx")
print("saved")

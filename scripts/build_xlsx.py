import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d=json.load(open("combos_data.json")); sc=json.load(open("scenarios.json"))
picks=d["picks"]; combos=d["combos"]; incl=d["coverage_incl"]; mc=d["mc"]
byid={p[0]:p for p in picks}
F="Arial"
hdr=Font(name=F,bold=True,color="FFFFFF",size=11)
hfill=PatternFill("solid",start_color="1F4E78")
sub=Font(name=F,bold=True,size=11)
base=Font(name=F,size=10)
center=Alignment(horizontal="center",vertical="center")
left=Alignment(horizontal="left",vertical="center")
thin=Side(style="thin",color="D0D0D0")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
tcolor={"AA":"C6EFCE","A":"FFF2CC","M":"FCE4D6"}

wb=Workbook()

# ---- Resumen ----
s=wb.active; s.title="Resumen"
s["A1"]="Mundial 2026 — Combinadas con cobertura"; s["A1"].font=Font(name=F,bold=True,size=16)
s["A2"]="12 combinadas de 20 partidos · S/20 cada una · S/240 total"; s["A2"].font=Font(name=F,size=11,italic=True)
r=4
s.cell(r,1,"Cómo funciona la cobertura").font=sub
notes=[
 "Los 41 picks se reparten en 12 combinadas distintas según tu nivel de confianza.",
 "AA (alta-alta) aparecen en 9 de 12 combinadas; A (alta) en 5-6; M (media) solo en 2.",
 "Cada partido queda FUERA de al menos 3 combinadas, así un fallo aislado no mata todo.",
 "OJO: cada combinada necesita que sus 20 patas acierten. La cobertura ayuda contra fallos sueltos, no contra varios a la vez.",
]
for n in notes:
    r+=1; c=s.cell(r,1,"• "+n); c.font=base; s.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
r+=2
s.cell(r,1,"Escenarios (simulación Monte Carlo, 200k rondas)").font=sub
r+=1
heads=["Escenario","P(que alguna gane)","Retorno promedio/ronda","Neto promedio/ronda"]
for j,h in enumerate(heads,1):
    c=s.cell(r,j,h); c.font=hdr; c.fill=hfill; c.alignment=center; c.border=border
for name,v in sc.items():
    r+=1
    vals=[name,f"{v['any']}%",f"S/ {v['ret']:,.0f}",f"S/ {v['net']:+,.0f}"]
    for j,val in enumerate(vals,1):
        c=s.cell(r,j,val); c.font=base; c.border=border
        c.alignment=left if j==1 else center
r+=2
s.cell(r,1,"Lectura honesta: si tus probabilidades reales son tan altas como crees, el conjunto es muy rentable.").font=Font(name=F,size=10,italic=True)
r+=1
s.cell(r,1,"Si se parecen a lo que cree la casa (cuotas), el resultado es prácticamente break-even o pérdida. La verdad suele estar en medio.").font=Font(name=F,size=10,italic=True)
r+=2
s.cell(r,1,f"Fallos esperados de tus 41 picks (escenario optimista): ~{mc['expected_failures_of_41']}").font=base
s.column_dimensions["A"].width=42
for col in "BCD": s.column_dimensions[col].width=22

# ---- Picks ----
s2=wb.create_sheet("Picks")
heads=["#","Pick","Partido","Cuota","Nivel","En # combinadas"]
for j,h in enumerate(heads,1):
    c=s2.cell(1,j,h); c.font=hdr; c.fill=hfill; c.alignment=center; c.border=border
for i,p in enumerate(picks,2):
    pid=p[0]
    row=[pid,p[1],p[2],p[3],p[4],f"{incl[str(pid)]}/12"]
    for j,val in enumerate(row,1):
        c=s2.cell(i,j,val); c.font=base; c.border=border
        c.alignment=center if j in(1,4,5,6) else left
    s2.cell(i,5).fill=PatternFill("solid",start_color=tcolor[p[4]])
    s2.cell(i,4).number_format="0.00"
s2.freeze_panes="A2"
for col,w in zip("ABCDEF",[5,26,30,9,8,16]): s2.column_dimensions[col].width=w

# ---- Combinadas ----
s3=wb.create_sheet("Combinadas")
row=1
for ci in combos:
    legs=ci["legs"]; cnum=ci["combo"]
    c=s3.cell(row,1,f"COMBINADA {cnum}  ·  S/20"); c.font=Font(name=F,bold=True,size=12,color="FFFFFF")
    c.fill=hfill
    for col in range(1,5):
        s3.cell(row,col).fill=hfill
    s3.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4)
    row+=1
    for h_i,h in enumerate(["#","Pick","Partido","Cuota"],1):
        cc=s3.cell(row,h_i,h); cc.font=sub; cc.border=border; cc.alignment=center
    row+=1
    first=row
    for pid in legs:
        p=byid[pid]
        s3.cell(row,1,pid).alignment=center
        s3.cell(row,2,p[1])
        s3.cell(row,3,p[2])
        oc=s3.cell(row,4,p[3]); oc.number_format="0.00"; oc.alignment=center
        for col in range(1,5):
            s3.cell(row,col).font=base; s3.cell(row,col).border=border
        s3.cell(row,1).fill=PatternFill("solid",start_color=tcolor[p[4]])
        row+=1
    last=row-1
    s3.cell(row,3,"Cuota total").font=sub; s3.cell(row,3).alignment=Alignment(horizontal="right")
    fc=s3.cell(row,4,f"=ROUND(PRODUCT(D{first}:D{last}),2)"); fc.font=Font(name=F,bold=True); fc.number_format="0.00"; fc.alignment=center
    cuota_cell=f"D{row}"; row+=1
    s3.cell(row,3,"Pago potencial (S/20)").font=sub; s3.cell(row,3).alignment=Alignment(horizontal="right")
    pc=s3.cell(row,4,f"={cuota_cell}*20"); pc.font=Font(name=F,bold=True,color="006100"); pc.number_format='"S/ "#,##0.00'; pc.alignment=center
    row+=2
for col,w in zip("ABCD",[5,26,30,12]): s3.column_dimensions[col].width=w

# ---- Cobertura ----
s4=wb.create_sheet("Cobertura")
s4.cell(1,1,"X = el pick está en esa combinada").font=Font(name=F,italic=True,size=10)
hr=2
s4.cell(hr,1,"#").font=hdr; s4.cell(hr,1,"#"); 
c=s4.cell(hr,1,"#"); c.font=hdr;c.fill=hfill;c.alignment=center;c.border=border
c=s4.cell(hr,2,"Pick"); c.font=hdr;c.fill=hfill;c.alignment=center;c.border=border
for k in range(12):
    c=s4.cell(hr,3+k,f"C{k+1}"); c.font=hdr;c.fill=hfill;c.alignment=center;c.border=border
combo_sets=[set(ci["legs"]) for ci in combos]
for i,p in enumerate(picks,hr+1):
    pid=p[0]
    c=s4.cell(i,1,pid);c.font=base;c.alignment=center;c.border=border
    c=s4.cell(i,2,p[1]);c.font=base;c.border=border
    c.fill=PatternFill("solid",start_color=tcolor[p[4]])
    for k in range(12):
        v="X" if pid in combo_sets[k] else ""
        cc=s4.cell(i,3+k,v);cc.font=base;cc.alignment=center;cc.border=border
        if v=="X": cc.fill=PatternFill("solid",start_color="DDEBF7")
s4.freeze_panes="C3"
s4.column_dimensions["A"].width=5; s4.column_dimensions["B"].width=24
for k in range(12): s4.column_dimensions[get_column_letter(3+k)].width=5

wb.save("Combinadas_Mundial_2026.xlsx")
print("saved")

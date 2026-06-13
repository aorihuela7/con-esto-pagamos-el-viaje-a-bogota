import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sfrac=json.load(open("sfrac.json"))
sfrac={int(k):v for k,v in sfrac.items()}
INFO={
1:("México gana",1.45,0.75),2:("Suiza gana",1.26,0.87),4:("Turquía gana",1.80,0.78),
5:("Alemania gana",1.05,0.98),6:("España gana",1.10,0.98),8:("Uruguay gana",1.50,0.6667),
9:("Francia gana",1.47,0.71),10:("Noruega gana",1.23,0.87),12:("Austria gana",1.34,0.82),
13:("Portugal gana",1.28,0.85),14:("Ghana 1X",1.29,0.80),15:("Colombia gana",1.42,0.92),
16:("Rep.Checa 1X",1.27,0.88),19:("Marruecos gana",2.05,0.82),20:("Brasil gana",1.06,0.99),
21:("Ecuador gana",1.27,0.95),23:("España gana",1.13,0.885),24:("Bélgica gana",1.47,0.73),
25:("Uruguay gana",1.47,0.76),26:("Egipto 2X",1.19,0.87),28:("Francia gana",1.12,0.96),
30:("Portugal gana",1.27,0.95),31:("Inglaterra gana",1.32,0.82),32:("Croacia gana",1.50,0.84),
33:("Colombia gana",1.52,0.75),34:("Brasil gana",1.40,0.92),35:("Marruecos gana",1.26,0.93),
36:("Corea 2X",1.27,0.80),37:("C.Marfil gana",1.34,0.84),38:("Senegal gana",1.47,0.73),
39:("Bélgica gana",1.30,0.86),40:("Inglaterra gana",1.20,0.95),41:("Argentina gana",1.20,0.91),
}
days=[("11/06",[1]),("13/06",[2,4]),("14/06",[5]),("15/06",[6]),("16/06",[9,10,12]),
("17/06",[13,14,15]),("18/06",[16]),("19/06",[19,20]),("20/06",[21]),("21/06",[24,25,26]),
("22/06",[28]),("23/06",[30,31,32,33]),("24/06",[34,35,36]),("25/06",[37]),("26/06",[38,39]),
("27/06",[40,41])]
# drop zero-stake picks (#8,#23) already excluded above

F="Arial"; WHITE="FFFFFF"; NAVY="1F4E78"; BLUE="0000FF"; GREEN="006100"; RED="C00000"
hdr=Font(name=F,bold=True,color=WHITE,size=10); hfill=PatternFill("solid",start_color=NAVY)
sub=Font(name=F,bold=True,size=10); base=Font(name=F,size=10)
inp=Font(name=F,size=10,color=BLUE,bold=True); inpfill=PatternFill("solid",start_color="FFF2CC")
center=Alignment(horizontal="center",vertical="center"); left=Alignment(horizontal="left",vertical="center")
right=Alignment(horizontal="right",vertical="center")
thin=Side(style="thin",color="C0C0C0"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
M2='"S/ "#,##0.00;[Red]"S/ -"#,##0.00'; MONEY='"S/ "#,##0.00'
dayfill=PatternFill("solid",start_color="D9E1F2"); totfill=PatternFill("solid",start_color="E2EFDA")

wb=Workbook()

# ---------- Resumen ----------
sr=wb.active; sr.title="Resumen"
sr["A1"]="Plan día a día — Mundial 2026 (banca rodante + Kelly fraccionado)"; sr["A1"].font=Font(name=F,bold=True,size=15)
sr["A2"]="Banca S/500 · ⅔ Kelly (tope 6%/apuesta) · bankear 40% de ganancias · formato mixto"; sr["A2"].font=Font(name=F,italic=True,size=10)
r=4
sr.cell(r,1,"QUÉ ESPERAR (simulación 40k)").font=sub; r+=1
heads=["Escenario","Banca final (mediana)","P(terminar en verde)","Rango P10–P90","Riesgo de ruina"]
for j,h in enumerate(heads,1):
    c=sr.cell(r,j,h); c.font=hdr; c.fill=hfill; c.alignment=center; c.border=border
data=[("Si tus probabilidades aciertan","S/ 676 (+35%)","92%","S/ 514 – 859","~0%"),
      ("Si el mercado tiene razón","S/ 488 (≈ plano)","46%","S/ 355 – 662","~0%")]
for row in data:
    r+=1
    for j,v in enumerate(row,1):
        c=sr.cell(r,j,v); c.font=base; c.border=border; c.alignment=left if j==1 else center
r+=2
for n in ["CÓMO USAR EL PLAN:",
 "1) La hoja 'Plan diario' es tu tracker. La banca inicial (amarillo) arriba ya está en S/500.",
 "2) Cada día están los partidos con VALOR y su stake en % de banca. El S/ a apostar se calcula solo según tu banca de ese día.",
 "3) Cuando termine cada partido, escribe 1 si ganó (o 0 si no) en la columna amarilla '¿Ganó?'. La banca se actualiza sola y rueda al día siguiente.",
 "4) 'Combo del día' es opcional (mezcla los partidos del día, stake chico) — el bono mixto.",
 "5) Para cashout/hedge en vivo, usa la hoja 'Cashout y Hedge'.",
 "REGLA DE ORO: solo apostamos partidos con edge>0. Los 8 sin valor quedaron fuera del plan."]:
    c=sr.cell(r,1,n); c.font=Font(name=F,bold=n.endswith(":"),size=10); sr.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); r+=1
for col,w in zip("ABCDE",[34,22,20,18,14]): sr.column_dimensions[col].width=w

# ---------- Plan diario ----------
s=wb.create_sheet("Plan diario")
s["A1"]="PLAN DIARIO — escribe 1/0 en '¿Ganó?' y todo se actualiza"; s["A1"].font=Font(name=F,bold=True,size=13)
s.merge_cells("A1:I1")
s["A2"]="Banca inicial:"; s["A2"].font=sub; s["A2"].alignment=right
bk0=s["B2"]; bk0.value=500; bk0.font=inp; bk0.fill=inpfill; bk0.number_format=MONEY; bk0.alignment=center
cols=["#","Pick / Apuesta","Cuota","Mi P","Edge","Stake %","S/ Apostado","¿Ganó? (1/0)","Retorno"]
widths=[5,26,7,7,7,8,12,12,12]
for col,w in zip("ABCDEFGHI",widths): s.column_dimensions[col].width=w

row=4
prev_close="B2"
boveda_all=[]
for di,(day,ids) in enumerate(days):
    # day banner
    c=s.cell(row,1,f"DÍA {day}"); c.font=Font(name=F,bold=True,size=11)
    for col in range(1,10): s.cell(row,col).fill=dayfill; s.cell(row,col).border=border
    s.cell(row,6,"Banca inicio →").font=Font(name=F,bold=True,size=10); s.cell(row,6).alignment=right
    binicio=f"G{row}"; cc=s.cell(row,7,f"={prev_close}"); cc.font=Font(name=F,bold=True); cc.number_format=MONEY; cc.alignment=center
    row+=1
    # header
    for j,h in enumerate(cols,1):
        c=s.cell(row,j,h); c.font=hdr; c.fill=hfill; c.alignment=center; c.border=border
    row+=1
    first=row
    apostado_cells=[]; retorno_cells=[]; gano_cells=[]
    for i in ids:
        pick,odds,p=INFO[i]; edge=p*odds-1; sf=sfrac[i]
        s.cell(row,1,i).alignment=center
        s.cell(row,2,pick)
        s.cell(row,3,odds).number_format="0.00"; s.cell(row,3).alignment=center
        s.cell(row,4,p).number_format="0%"; s.cell(row,4).alignment=center
        s.cell(row,5,edge).number_format="+0.0%;-0.0%"; s.cell(row,5).alignment=center
        s.cell(row,6,sf).number_format="0.0%"; s.cell(row,6).alignment=center
        ap=s.cell(row,7,f'=IF($H{row}="",0,$F{row}*{binicio})'); ap.number_format=MONEY; ap.alignment=center
        g=s.cell(row,8,None); g.font=inp; g.fill=inpfill; g.alignment=center
        rt=s.cell(row,9,f"=IF($H{row}=1,$G{row}*$C{row},0)"); rt.number_format=MONEY; rt.alignment=center
        for col in range(1,10): s.cell(row,col).border=border
        for col in [1,2,3,4,5,6]: s.cell(row,col).font=base
        s.cell(row,7).font=base
        apostado_cells.append(f"G{row}"); retorno_cells.append(f"I{row}"); gano_cells.append(f"H{row}")
        row+=1
    # combo del día (optional) if >=2
    if len(ids)>=2:
        co=1.0
        for i in ids: co*=INFO[i][1]
        s.cell(row,1,"C").alignment=center
        s.cell(row,2,f"Combo del día ({'+'.join('#'+str(i) for i in ids)})")
        s.cell(row,3,round(co,2)).number_format="0.00"; s.cell(row,3).alignment=center
        s.cell(row,6,0.02).number_format="0.0%"; s.cell(row,6).alignment=center
        ap=s.cell(row,7,f'=IF($H{row}="",0,$F{row}*{binicio})'); ap.number_format=MONEY; ap.alignment=center
        # combo gana si todos los del día ganaron (en blanco hasta que se jueguen todos)
        cond="+".join(gano_cells); cnt=",".join(gano_cells)
        g=s.cell(row,8,f"=IF(COUNT({cnt})<{len(ids)},\"\",IF(({cond})={len(ids)},1,0))"); g.alignment=center; g.font=Font(name=F,size=10,italic=True)
        rt=s.cell(row,9,f"=IF($H{row}=1,$G{row}*$C{row},0)"); rt.number_format=MONEY; rt.alignment=center
        for col in range(1,10): s.cell(row,col).border=border; 
        for col in [1,2,3,6,7]: s.cell(row,col).font=base if col!=2 else Font(name=F,size=10,italic=True)
        apostado_cells.append(f"G{row}"); retorno_cells.append(f"I{row}")
        row+=1
    # day totals
    ap_sum="+".join(apostado_cells); rt_sum="+".join(retorno_cells)
    s.cell(row,2,"Apostado / Retorno / Neto del día").font=sub; s.cell(row,2).alignment=right
    ca=s.cell(row,7,f"={ap_sum}"); ca.number_format=MONEY; ca.alignment=center; ca.font=sub
    cr=s.cell(row,8,f"={rt_sum}"); cr.number_format=MONEY; cr.alignment=center; cr.font=sub
    neto_cell=f"I{row}"; cn=s.cell(row,9,f"=({rt_sum})-({ap_sum})"); cn.number_format=M2; cn.alignment=center; cn.font=Font(name=F,bold=True)
    for col in range(1,10): s.cell(row,col).fill=totfill; s.cell(row,col).border=border
    row+=1
    # bank close
    s.cell(row,2,"A bóveda (40% si ganó) / Banca cierre").font=sub; s.cell(row,2).alignment=right
    boveda=f"H{row}"; cb=s.cell(row,8,f"=IF({neto_cell}>0,0.4*{neto_cell},0)"); cb.number_format=MONEY; cb.alignment=center; cb.font=base
    boveda_all.append(boveda)
    close=f"I{row}"; cc=s.cell(row,9,f"={binicio}+{neto_cell}-{boveda}"); cc.number_format=MONEY; cc.alignment=center; cc.font=Font(name=F,bold=True,color=GREEN)
    s.cell(row,6,"Banca cierre →").font=sub; s.cell(row,6).alignment=right
    for col in range(1,10): s.cell(row,col).fill=totfill; s.cell(row,col).border=border
    prev_close=close
    row+=2

# Final summary
boveda_sum="+".join(boveda_all)
s.cell(row,2,"Banca rodante final").font=Font(name=F,bold=True,size=11); s.cell(row,2).alignment=right
cc=s.cell(row,9,f"={prev_close}"); cc.number_format=MONEY; cc.alignment=center; cc.font=Font(name=F,bold=True)
for col in range(1,10): s.cell(row,col).fill=totfill; s.cell(row,col).border=border
row+=1
s.cell(row,2,"Bóveda acumulada (asegurado)").font=Font(name=F,bold=True,size=11); s.cell(row,2).alignment=right
cb=s.cell(row,9,f"={boveda_sum}"); cb.number_format=MONEY; cb.alignment=center; cb.font=Font(name=F,bold=True,color=GREEN)
for col in range(1,10): s.cell(row,col).fill=totfill; s.cell(row,col).border=border
row+=1
s.cell(row,2,"RIQUEZA TOTAL (banca + bóveda)").font=Font(name=F,bold=True,size=12); s.cell(row,2).alignment=right
ct=s.cell(row,9,f"={prev_close}+{boveda_sum}"); ct.number_format=MONEY; ct.alignment=center; ct.font=Font(name=F,bold=True,color=GREEN,size=12)
for col in range(1,10): s.cell(row,col).fill=PatternFill("solid",start_color="C6EFCE"); s.cell(row,col).border=border
row+=2
s.cell(row,2,"Nota: riqueza total = banca rodante final + lo que mandaste a la bóveda en el camino.").font=Font(name=F,italic=True,size=9)

# ---------- Cashout y Hedge (reuse) ----------
ch=wb.create_sheet("Cashout y Hedge")
ch.cell(1,1,"Gestión en vivo: Cashout y Cobertura (Hedge)").font=Font(name=F,bold=True,size=13)
r=3
ch.cell(r,1,"REGLA DE CASHOUT").font=sub; r+=1
for n in ["Valor esperado de seguir = (prob. de las patas restantes) × (pago potencial).",
 "Cierra cuando el cashout ofrecido ≥ ese valor, o cuando ya asegura tu objetivo y no quieres arriesgar.",
 "Con banca rodante conviene cerrar antes: proteger el capital que rueda al día siguiente vale más que el último 10%."]:
    c=ch.cell(r,1,"• "+n); c.font=base; ch.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); r+=1
r+=1
ch.cell(r,1,"CALCULADORA DE HEDGE (última pata viva)").font=sub; r+=1
calc=[("Pago potencial si gana (W, S/)",1000,MONEY),("Cuota del resultado CONTRARIO (o)",2.50,"0.00"),("Ya apostado (S, S/)",20,MONEY)]
addr={}
for lab,val,fmt in calc:
    ch.cell(r,1,lab).font=base; cc=ch.cell(r,2,val); cc.font=inp; cc.fill=inpfill; cc.number_format=fmt; cc.alignment=center; addr[lab]=f"B{r}"; r+=1
W,O,S=addr[calc[0][0]],addr[calc[1][0]],addr[calc[2][0]]
for lab,fml in [("Apostar al contrario (hedge=W/o)",f"={W}/{O}"),("Ganancia GARANTIZADA",f"={W}*(1-1/{O})-{S}")]:
    ch.cell(r,1,lab).font=sub; cc=ch.cell(r,2,fml); cc.font=Font(name=F,bold=True,color=GREEN); cc.number_format=M2; cc.alignment=center; r+=1
ch.column_dimensions["A"].width=40; ch.column_dimensions["B"].width=16

wb.save("Plan_Diario_Mundial.xlsx")
print("saved")

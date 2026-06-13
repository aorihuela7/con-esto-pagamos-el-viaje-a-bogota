from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
AR="Arial"; NAVY="1F3864"; GREEN="1F6E43"; LBLUE="DDEBF7"
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=load_workbook("Sistema_Mundial_2026_COMPLETO.xlsx")
p=wb["Portada"]
# mover recordatorio (fila 22) hacia abajo: lo reescribimos mas abajo
old=p.cell(22,2).value
p.cell(22,2).value=None
r=22
p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
c=p.cell(r,2,"4 · COMBINADA SEGURA (BETSSON)"); c.font=f(12,True,"FFFFFF"); c.fill=fill(GREEN); c.alignment=lft; p.row_dimensions[r].height=24; r+=1
items=[
 ("Segura · Resumen","12 combinadas de S/60 con el reembolso de Betsson (si falla 1 pata) + diversificacion."),
 ("Segura · Detalle","Los 5 partidos de cada una de las 12 combinadas seguras."),
 ("Segura · Mapa","Mapa de diversificacion de las combinadas seguras (✓ por combinada)."),
]
for name,desc in items:
    cl=p.cell(r,2,name); cl.font=f(10,True,NAVY); cl.alignment=lft; cl.border=bd; cl.fill=fill(LBLUE)
    cd=p.cell(r,3,desc); cd.font=f(9); cd.alignment=lft; cd.border=bd
    p.row_dimensions[r].height=24; r+=1
r+=1
p.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
c=p.cell(r,2,old); c.font=f(9,False,"808080"); c.alignment=lft; p.row_dimensions[r].height=46
wb.save("Sistema_Mundial_2026_COMPLETO.xlsx")
print("portada actualizada; hojas:",len(wb.sheetnames))

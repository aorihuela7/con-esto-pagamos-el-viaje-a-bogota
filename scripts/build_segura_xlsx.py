import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

d=json.load(open("segura.json"))
combos=d["combos"]; picks=d["picks"]; use=d["use"]; wcount=d["wcount"]
AR="Arial"
NAVY="1F3864"; BLUE="2E5496"; GREEN="1F6E43"; LGREEN="C6EFCE"; MGREEN="E2EFDA"
GOLD="BF9000"; LGOLD="FFF2CC"; LGREY="F2F2F2"; LBLUE="DDEBF7"; RED="9C0006"; LRED="FFC7CE"
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=load_workbook("Sistema_Mundial_2026_COMPLETO.xlsx")
for nm in ["Segura · Resumen","Segura · Detalle","Segura · Mapa"]:
    if nm in wb.sheetnames: del wb[nm]

# ================= HOJA 1: RESUMEN =================
s=wb.create_sheet("Segura · Resumen"); s.sheet_view.showGridLines=False
widths=[2,6,7,11,11,11,12,13,12,12,16,2]
for col,w in zip("ABCDEFGHIJKL",widths): s.column_dimensions[col].width=w
s.merge_cells("B2:K2")
c=s["B2"]; c.value="COMBINADA SEGURA · BETSSON  ·  12 combinadas de S/60"; c.font=f(15,True,"FFFFFF"); c.fill=fill(GREEN); c.alignment=ctr; s.row_dimensions[2].height=28
s.merge_cells("B3:K3")
c=s["B3"]; c.value=("Promo: combinada de 5+ patas, SOLO ganador directo (1X2) con cuota >=1.40 c/u y total >=5.30. "
 "Si falla EXACTAMENTE una pata (4 ganan, 1 pierde) te DEVUELVEN lo apostado hasta S/60. "
 "Apuesta S/60 = reembolso cubre el 100%. Maximo 3 reembolsos por dia (ventana 17:00-16:59), segun el partido que termina mas tarde.")
c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=46

r=5
heads=["#","Patas","Cuota total","Paga S/","Ganancia S/","P(5 ganan)","P(1 falla→reembolso)","P(2+ →pierdes)","EV con seguro","Ventana reembolso","Partido mas tardio"]
for k,h in enumerate(heads):
    cc=s.cell(r,2+k,h); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(GREEN); cc.alignment=ctr; cc.border=bd
s.row_dimensions[r].height=30
r+=1; datstart=r
for k,cb in enumerate(combos,1):
    row=[k,len(cb["legs"]),round(cb["odds"],2),cb["payout"],cb["profit"],
         cb["p_all"],cb["p1"],cb["p2plus"],round(cb["ev"]),cb["win"],
         f'{cb["last"]} ({cb["last_end"]})']
    fmts=[None,None,"0.00",'"S/ "#,##0','"S/ "#,##0',"0%","0%","0%",'"S/ "+#,##0;"S/ "-#,##0',None,None]
    for j,(val,fm) in enumerate(zip(row,fmts)):
        cc=s.cell(r,2+j,val); cc.border=bd; cc.alignment=(lft if j==10 else ctr)
        cc.font=f(9, j in(0,4,8))
        if fm: cc.number_format=fm
        if j==0: cc.fill=fill(MGREEN)
        if j==7: cc.font=f(9,False,RED if val>0.30 else "000000")     # P2+ rojo si alto
        if j==8: cc.font=f(9,True,"006100" if val>=0 else RED)
    s.row_dimensions[r].height=22; r+=1
datend=r-1
# fila totales
s.cell(r,2,"TOTAL").font=f(10,True); s.cell(r,2).fill=fill(LGREEN); s.cell(r,2).border=bd; s.cell(r,2).alignment=ctr
s.merge_cells(start_row=r,start_column=3,end_row=r,end_column=4)
cc=s.cell(r,3,f"Costo: S/{int(d['cost'])}"); cc.font=f(10,True); cc.alignment=ctr; cc.fill=fill(LGREEN); cc.border=bd; s.cell(r,4).fill=fill(LGREEN); s.cell(r,4).border=bd
s.merge_cells(start_row=r,start_column=5,end_row=r,end_column=9)
cc=s.cell(r,5,"EV total con seguro →"); cc.font=f(10,True); cc.alignment=Alignment(horizontal="right",vertical="center"); cc.fill=fill(LGREEN); cc.border=bd
for cx in range(6,10): s.cell(r,cx).fill=fill(LGREEN); s.cell(r,cx).border=bd
cc=s.cell(r,10,round(d['agg_ev'])); cc.number_format='"S/ "+#,##0'; cc.font=f(11,True,"006100"); cc.alignment=ctr; cc.fill=fill(LGREEN); cc.border=bd
s.merge_cells(start_row=r,start_column=11,end_row=r,end_column=12); s.cell(r,11).fill=fill(LGREEN); s.cell(r,11).border=bd; s.cell(r,12).border=bd
s.row_dimensions[r].height=22; r+=2

# bloque diversificacion
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=11)
c=s.cell(r,2,"DIVERSIFICACION  ·  que pasa si fallan partidos"); c.font=f(11,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=lft; s.row_dimensions[r].height=22; r+=1
div=[
 (f"Min. partidos que deben fallar para tumbar las 12 combinadas: {d['mh']}  ({', '.join(d['mhset'])})",
  "Si falla 1 o 2 partidos cualesquiera, SIEMPRE sobrevive al menos una combinada."),
 (f"Cobertura si falla 1 partido: {d['cov1'][0]}/{d['cov1'][1]} = 100%",
  "Pase lo que pase con un solo resultado, te quedan combinadas vivas."),
 (f"Cobertura si fallan 2 partidos: {d['cov2'][0]}/{d['cov2'][1]} = 100%",
  "Cualquier par de fallos deja combinadas en pie."),
 ("Ningun pick se repite en mas de 4 de las 12 combinadas.",
  "Asi un mal resultado no contagia a todas."),
]
for a,b in div:
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    ca=s.cell(r,2,a); ca.font=f(9,True,NAVY); ca.alignment=lft; ca.border=bd; ca.fill=fill(LBLUE)
    s.merge_cells(start_row=r,start_column=6,end_row=r,end_column=11)
    cbx=s.cell(r,6,b); cbx.font=f(9); cbx.alignment=lft; cbx.border=bd
    s.row_dimensions[r].height=24; r+=1
r+=1

# reparto por ventana
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=11)
c=s.cell(r,2,"REPARTO POR VENTANA DE REEMBOLSO (max. 3 cupones por dia)"); c.font=f(11,True,"FFFFFF"); c.fill=fill(GOLD); c.alignment=lft; s.row_dimensions[r].height=20; r+=1
s.cell(r,2,"Ventana (17:00 → 16:59)").font=f(9,True); s.cell(r,2).border=bd; s.cell(r,2).fill=fill(LGOLD); s.cell(r,2).alignment=ctr
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
s.cell(r,3).fill=fill(LGOLD); s.cell(r,3).border=bd; s.cell(r,4).fill=fill(LGOLD); s.cell(r,4).border=bd
s.cell(r,5,"Combinadas que cierran").font=f(9,True); s.cell(r,5).border=bd; s.cell(r,5).fill=fill(LGOLD); s.cell(r,5).alignment=ctr
s.merge_cells(start_row=r,start_column=5,end_row=r,end_column=7)
s.cell(r,6).fill=fill(LGOLD); s.cell(r,6).border=bd; s.cell(r,7).fill=fill(LGOLD); s.cell(r,7).border=bd
s.cell(r,8,"Reembolsos posibles").font=f(9,True); s.cell(r,8).border=bd; s.cell(r,8).fill=fill(LGOLD); s.cell(r,8).alignment=ctr
s.merge_cells(start_row=r,start_column=8,end_row=r,end_column=11)
for cx in range(9,12): s.cell(r,cx).fill=fill(LGOLD); s.cell(r,cx).border=bd
r+=1
for win,cnt in sorted(wcount.items()):
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
    cc=s.cell(r,2,win); cc.font=f(9); cc.alignment=ctr; cc.border=bd
    for cx in (3,4): s.cell(r,cx).border=bd
    s.merge_cells(start_row=r,start_column=5,end_row=r,end_column=7)
    cc=s.cell(r,5,cnt); cc.font=f(9,True); cc.alignment=ctr; cc.border=bd
    for cx in (6,7): s.cell(r,cx).border=bd
    s.merge_cells(start_row=r,start_column=8,end_row=r,end_column=11)
    cc=s.cell(r,8,f"hasta S/{min(cnt,3)*60} ({min(cnt,3)} cupon(es))"); cc.font=f(9); cc.alignment=ctr; cc.border=bd
    for cx in range(9,12): s.cell(r,cx).border=bd
    r+=1
r+=1
for t in [
 "Lee el EV con seguro: lo que esperas ganar por combinada contando el reembolso si falla 1 pata. Todas son +EV con tus probabilidades.",
 "Regla de oro: apuesta EXACTAMENTE S/60 por combinada (iguala el tope de reembolso) y NO uses cashout (te descalifica del reembolso).",
 "El seguro solo paga si falla UNA pata. Por eso conviene 5 patas (no mas) y las de mayor seguridad: menos chance de que fallen 2+.",
 "Las P() usan TUS probabilidades (columna Mi prob.), no las de la casa. Si las cambias, cambian estos numeros.",
]:
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=11)
    cc=s.cell(r,2,"•  "+t); cc.font=f(9); cc.alignment=lft; s.row_dimensions[r].height=24; r+=1

# ================= HOJA 2: DETALLE =================
s2=wb.create_sheet("Segura · Detalle"); s2.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[2,8,30,9,9,2]): s2.column_dimensions[col].width=w
s2.merge_cells("B2:E2")
c=s2["B2"]; c.value="DETALLE DE CADA COMBINADA SEGURA"; c.font=f(14,True,"FFFFFF"); c.fill=fill(GREEN); c.alignment=ctr; s2.row_dimensions[2].height=24
r=4
for k,cb in enumerate(combos,1):
    s2.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    head=(f"COMBINADA #{k}   ·   cuota {cb['odds']:.2f}   ·   paga S/{cb['payout']:,}   ·   "
          f"reembolso en ventana {cb['win']}")
    cc=s2.cell(r,2,head); cc.font=f(10,True,"FFFFFF"); cc.fill=fill(BLUE); cc.alignment=lft; s2.row_dimensions[r].height=20; r+=1
    for col,h in zip(range(2,6),["Pick","Partido","Cuota","Mi prob."]):
        cc=s2.cell(r,col,h); cc.font=f(9,True); cc.fill=fill(LGREEN); cc.alignment=ctr; cc.border=bd
    r+=1
    for i in cb["legs"]:
        eq,part,od,pr,ko=picks[str(i)]
        cc=s2.cell(r,2,eq); cc.font=f(9,True,NAVY); cc.alignment=lft; cc.border=bd
        cc=s2.cell(r,3,f"{part}  ({ko})"); cc.font=f(9); cc.alignment=lft; cc.border=bd
        cc=s2.cell(r,4,od); cc.number_format="0.00"; cc.font=f(9); cc.alignment=ctr; cc.border=bd
        cc=s2.cell(r,5,pr); cc.number_format="0%"; cc.font=f(9,False,"0000FF"); cc.alignment=ctr; cc.border=bd; cc.fill=fill(LGOLD)
        r+=1
    # fila combinada
    cc=s2.cell(r,2,"COMBINADA"); cc.font=f(9,True); cc.alignment=lft; cc.border=bd; cc.fill=fill(MGREEN)
    cc=s2.cell(r,3,f"5 patas → paga S/{cb['payout']:,} (gana S/{cb['profit']:,})"); cc.font=f(9,True); cc.alignment=lft; cc.border=bd; cc.fill=fill(MGREEN)
    cc=s2.cell(r,4,round(cb['odds'],2)); cc.number_format="0.00"; cc.font=f(9,True); cc.alignment=ctr; cc.border=bd; cc.fill=fill(MGREEN)
    cc=s2.cell(r,5,cb['p_all']); cc.number_format="0%"; cc.font=f(9,True,"006100"); cc.alignment=ctr; cc.border=bd; cc.fill=fill(MGREEN)
    r+=2

# ================= HOJA 3: MAPA =================
s3=wb.create_sheet("Segura · Mapa"); s3.sheet_view.showGridLines=False
s3.column_dimensions["A"].width=2
s3.column_dimensions["B"].width=14
s3.column_dimensions["C"].width=26
for col in "DEFGHIJKLMNO": s3.column_dimensions[col].width=6.6
s3.column_dimensions["P"].width=7
ncol=len(combos)
last_combo_col=3+ncol  # D..(3+12)=15 -> col O
s3.merge_cells(start_row=2,start_column=2,end_row=2,end_column=last_combo_col+1)
c=s3.cell(2,2,"MAPA DE DIVERSIFICACION  ·  picks (filas) x combinadas (columnas)"); c.font=f(13,True,"FFFFFF"); c.fill=fill(GREEN); c.alignment=ctr; s3.row_dimensions[2].height=24
r=4
s3.cell(r,2,"Pick").font=f(9,True,"FFFFFF"); s3.cell(r,2).fill=fill(NAVY); s3.cell(r,2).alignment=ctr; s3.cell(r,2).border=bd
s3.cell(r,3,"Partido").font=f(9,True,"FFFFFF"); s3.cell(r,3).fill=fill(NAVY); s3.cell(r,3).alignment=ctr; s3.cell(r,3).border=bd
for k in range(ncol):
    cc=s3.cell(r,4+k,f"C{k+1}"); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(NAVY); cc.alignment=ctr; cc.border=bd
cc=s3.cell(r,4+ncol,"Veces"); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(NAVY); cc.alignment=ctr; cc.border=bd
r+=1
# map leg->combos
legmap={int(i):[] for i in picks}
for k,cb in enumerate(combos):
    for i in cb["legs"]: legmap[i].append(k)
for idx,i in enumerate(sorted(legmap, key=lambda x:int(x))):
    eq,part,od,pr,ko=picks[str(i)]
    shade=fill(LGREY if idx%2 else "FFFFFF")
    cc=s3.cell(r,2,f"{eq} {od:.2f}"); cc.font=f(8,True,NAVY); cc.alignment=lft; cc.border=bd; cc.fill=shade
    cc=s3.cell(r,3,part); cc.font=f(8); cc.alignment=lft; cc.border=bd; cc.fill=shade
    for k in range(ncol):
        cc=s3.cell(r,4+k)
        if k in legmap[i]:
            cc.value="✓"; cc.font=f(10,True,"006100"); cc.fill=fill(LGREEN)
        else:
            cc.fill=shade
        cc.alignment=ctr; cc.border=bd
    cc=s3.cell(r,4+ncol,len(legmap[i])); cc.font=f(9,True); cc.alignment=ctr; cc.border=bd; cc.fill=shade
    r+=1
# fila patas por combinada
cc=s3.cell(r,2,"Patas"); cc.font=f(9,True); cc.alignment=ctr; cc.border=bd; cc.fill=fill(MGREEN)
cc=s3.cell(r,3,"(cada combinada = 5)"); cc.font=f(8); cc.alignment=lft; cc.border=bd; cc.fill=fill(MGREEN)
for k,cb in enumerate(combos):
    cc=s3.cell(r,4+k,len(cb["legs"])); cc.font=f(9,True); cc.alignment=ctr; cc.border=bd; cc.fill=fill(MGREEN)
s3.cell(r,4+ncol).fill=fill(MGREEN); s3.cell(r,4+ncol).border=bd
r+=1
# fila: ganancia neta apostando S/60 por combinada
cc=s3.cell(r,2,"Gana (S/60)"); cc.font=f(9,True,"006100"); cc.alignment=ctr; cc.border=bd; cc.fill=fill(LGREEN)
cc=s3.cell(r,3,"ganancia neta si la combinada gana"); cc.font=f(8); cc.alignment=lft; cc.border=bd; cc.fill=fill(LGREEN)
for k,cb in enumerate(combos):
    cc=s3.cell(r,4+k,cb["profit"]); cc.number_format='#,##0'; cc.font=f(8,True,"006100"); cc.alignment=ctr; cc.border=bd; cc.fill=fill(LGREEN)
cc=s3.cell(r,4+ncol,sum(cb["profit"] for cb in combos)); cc.number_format='"S/"#,##0'; cc.font=f(8,True,"006100"); cc.alignment=ctr; cc.border=bd; cc.fill=fill(LGREEN)
r+=2
s3.merge_cells(start_row=r,start_column=2,end_row=r,end_column=last_combo_col+1)
cc=s3.cell(r,2,f"Min. hitting set = {d['mh']}: hacen falta al menos {d['mh']} partidos fallando (de combos distintos) para que NINGUNA combinada sobreviva. Con 1 o 2 fallos siempre te queda alguna viva.")
cc.font=f(9,False,"808080"); cc.alignment=lft; s3.row_dimensions[r].height=30

wb.save("Sistema_Mundial_2026_COMPLETO.xlsx")
print("hojas:",[n for n in wb.sheetnames if n.startswith("Segura")])
print("total hojas:",len(wb.sheetnames))

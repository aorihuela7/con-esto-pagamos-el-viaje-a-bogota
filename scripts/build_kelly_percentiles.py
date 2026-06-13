import json
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, Series
np.random.seed(11)

d=json.load(open("cmp.json"))
od=np.array(d["od"]); pdp=np.array(d["pd"]); nd=len(od)
B0=100.0; SIM=200000
fracs=[("½ Kelly (conservador)",0.5,"006100","C6EFCE"),
       ("⅔ Kelly (medio)",2/3,"7F6000","FFF2CC"),
       ("Kelly 100% (agresivo)",1.0,"843C0C","FCE4D6")]

def kf(o,p,frac): return np.clip((p*o-1)/(o-1)*frac,0,1)

sim={}
for name,frac,_,_ in fracs:
    bank=np.full(SIM,B0)
    for i in range(nd):
        bet=bank*kf(od[i],pdp[i],frac)
        win=np.random.rand(SIM)<pdp[i]
        bank=bank+np.where(win,bet*(od[i]-1),-bet)
    sim[name]=bank

deciles=[10,20,30,40,50,60,70,80,90]
stats={}
for name,frac,_,_ in fracs:
    b=sim[name]
    stats[name]={"pct":{q:float(np.percentile(b,q)) for q in deciles},
                 "p_prof":float((b>B0).mean()*100),
                 "p_ruina":float((b<B0*0.5).mean()*100),
                 "mean":float(b.mean())}

AR="Arial"; NAVY="1F3864"; BLUE="2E5496"; LGREY="F2F2F2"; LBLUE="DDEBF7"
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=load_workbook("Sistema_Mundial_2026_COMPLETO.xlsx")
name="Kelly · Percentiles"
if name in wb.sheetnames: del wb[name]
s=wb.create_sheet(name); s.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGH",[2,16,13,13,13,13,13,13]): s.column_dimensions[col].width=w

s.merge_cells("B2:H2")
c=s["B2"]; c.value="DISTRIBUCIÓN POR PERCENTILES  ·  los 3 Kelly (combinada por día)"; c.font=f(15,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s.row_dimensions[2].height=28
s.merge_cells("B3:H3")
c=s["B3"]; c.value=(f"Banca inicial S/100, {nd} días apostando una combinada cada día con la fracción de Kelly indicada. "
 f"{SIM:,} mundiales simulados con TUS probabilidades. Cada percentil = % de casos que terminaron por DEBAJO de esa banca.")
c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=34

# header: two cols per kelly (Banca, Ganancia)
r=5
s.cell(r,2,"Percentil").font=f(9,True,"FFFFFF"); s.cell(r,2).fill=fill(BLUE); s.cell(r,2).alignment=ctr; s.cell(r,2).border=bd
s.merge_cells(start_row=r,start_column=2,end_row=r+1,end_column=2)
col=3
for name_k,frac,acol,light in fracs:
    s.merge_cells(start_row=r,start_column=col,end_row=r,end_column=col+1)
    cc=s.cell(r,col,name_k); cc.font=f(10,True,acol); cc.fill=fill(light); cc.alignment=ctr; cc.border=bd
    s.cell(r,col+1).fill=fill(light); s.cell(r,col+1).border=bd
    s.cell(r+1,col,"Banca").font=f(8,True); s.cell(r+1,col).alignment=ctr; s.cell(r+1,col).border=bd; s.cell(r+1,col).fill=fill(light)
    s.cell(r+1,col+1,"Ganancia").font=f(8,True); s.cell(r+1,col+1).alignment=ctr; s.cell(r+1,col+1).border=bd; s.cell(r+1,col+1).fill=fill(light)
    col+=2
r+=2
labelp={10:"P10 (peor 10%)",20:"P20",30:"P30",40:"P40",50:"P50 (mediana)",60:"P60",70:"P70",80:"P80",90:"P90 (mejor 10%)"}
datstart=r
for q in deciles:
    cl=s.cell(r,2,labelp[q]); cl.font=f(9,True if q in(10,50,90) else False); cl.alignment=lft; cl.border=bd
    cl.fill=fill(LBLUE if q==50 else (LGREY if (q//10)%2 else "FFFFFF"))
    col=3
    for name_k,frac,acol,light in fracs:
        b=stats[name_k]["pct"][q]
        cb=s.cell(r,col,round(b)); cb.number_format='"S/ "#,##0'; cb.font=f(10,True if q==50 else False); cb.alignment=ctr; cb.border=bd
        cg=s.cell(r,col+1,round(b-B0)); cg.number_format='"S/ "+#,##0;"S/ "-#,##0'; cg.alignment=ctr; cg.border=bd
        cg.font=f(9,False,"006100" if b>=B0 else "9C0006")
        col+=2
    r+=1
datend=r-1
# footer stats
for lab,key,fmt in [("Prob. de terminar ganando","p_prof","{:.1f} %"),
                    ("Prob. de perder la mitad+","p_ruina","{:.1f} %"),
                    ("Promedio (media)","mean","S/ {:.0f}")]:
    cl=s.cell(r,2,lab); cl.font=f(9,True); cl.alignment=lft; cl.border=bd; cl.fill=fill("FFFFFF")
    col=3
    for name_k,frac,acol,light in fracs:
        s.merge_cells(start_row=r,start_column=col,end_row=r,end_column=col+1)
        cc=s.cell(r,col,fmt.format(stats[name_k][key])); cc.font=f(10,True,acol); cc.alignment=ctr; cc.border=bd; cc.fill=fill(light)
        s.cell(r,col+1).border=bd; s.cell(r,col+1).fill=fill(light)
        col+=2
    r+=1
r+=1

# chart: deciles x-axis, 3 lines (banca)
ch=LineChart(); ch.title="Banca final por percentil — los 3 Kelly"; ch.height=9; ch.width=20
ch.y_axis.title="Banca (S/)"; ch.x_axis.title="Percentil"
cats=Reference(s,min_col=2,min_row=datstart,max_row=datend)
for k,(name_k,frac,acol,light) in enumerate(fracs):
    ref=Reference(s,min_col=3+2*k,min_row=datstart-1,max_row=datend)  # include header row label? header is at datstart-1? header banca row is r=5+1
    # banca columns are 3,5,7 ; header 'Banca' at row datstart-1
    ser=Series(Reference(s,min_col=3+2*k,min_row=datstart,max_row=datend),title=name_k)
    ch.series.append(ser)
ch.set_categories(cats)
s.add_chart(ch,f"B{r}")
r+=20

# notes
notes=[
 "Cómo leerlo: el P50 (mediana) es el resultado típico. El P10 es el mal escenario (solo 10% terminó peor); el P90 el buen escenario (solo 10% terminó mejor).",
 "½ Kelly: la curva más plana y segura. Casi nunca pierdes, pero el techo (P90) es el más bajo.",
 "⅔ Kelly: sube toda la curva sin meter riesgo de quiebra. Buen punto medio para tu perfil.",
 "Kelly 100%: la curva más empinada — mayor P90, pero el P10 baja y aparece algo de riesgo de perder la mitad de la banca.",
 "La distribución NO es normal (campana): es asimétrica a la derecha (lognormal). Por eso la media es mayor que la mediana — unos pocos casos muy buenos jalan el promedio.",
]
for t in notes:
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=8)
    c=s.cell(r,2,"•  "+t); c.font=f(9); c.alignment=lft; s.row_dimensions[r].height=26; r+=1

wb.save("Sistema_Mundial_2026_COMPLETO.xlsx")
print("ok percentiles. dias=",nd)
for name_k,frac,_,_ in fracs:
    p=stats[name_k]["pct"]
    print(f"{name_k:24s} P10 S/{p[10]:6.0f}  P50 S/{p[50]:6.0f}  P90 S/{p[90]:7.0f}  ganar {stats[name_k]['p_prof']:.0f}%")

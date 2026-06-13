import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

d=json.load(open("cmp2.json")); R=d["res"]; nbet=d["nbet"]; fmax=d["fmax"]
H,Tw,Fu=R["half"],R["twothirds"],R["full"]

AR="Arial"; NAVY="1F3864"; BLUE="2E5496"; GOLD="BF9000"
LGREY="F2F2F2"; LBLUE="DDEBF7"; LGOLD="FFF2CC"; GREEN="C6EFCE"; RED="F4CCCC"; ORANGE="FCE4D6"
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=load_workbook("Escalera_Mundial_41.xlsx")
if "Comparación Kelly" in wb.sheetnames: del wb["Comparación Kelly"]
s=wb.create_sheet("Comparación Kelly")
s.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[2,34,15,15,15,3]): s.column_dimensions[col].width=w

s.merge_cells("B2:E2")
c=s["B2"]; c.value="¿QUÉ FRACCIÓN DE KELLY?  —  Escalera partido a partido"; c.font=f(15,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s.row_dimensions[2].height=28
s.merge_cells("B3:E3")
c=s["B3"]; c.value=(f"Apuestas los {nbet} partidos con valor, uno por uno en orden de fecha, arriesgando una fracción de Kelly en cada uno. "
 "200.000 mundiales simulados con TUS probabilidades, banca inicial S/100.")
c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=32

r=5
hdr=["Métrica","½ Kelly (conservador)","⅔ Kelly (medio)","Kelly 100% (agresivo)"]
fills=[BLUE,GREEN,LGOLD,ORANGE]
for j,h in enumerate(hdr):
    cc=s.cell(r,2+j,h); cc.font=f(10,True,"FFFFFF" if j==0 else "000000"); cc.fill=fill(fills[j]); cc.alignment=ctr; cc.border=bd
s.row_dimensions[r].height=30
r+=1
def money(x): return f"S/ {x:.0f}"
rows=[
 ("Banca típica (mediana)", money(H['med']), money(Tw['med']), money(Fu['med'])),
 ("Prob. de terminar ganando", f"{H['p_prof']:.1f} %", f"{Tw['p_prof']:.1f} %", f"{Fu['p_prof']:.1f} %"),
 ("Prob. de quebrar (perder 90%+)", f"{H['p_ruina']:.1f} %", f"{Tw['p_ruina']:.1f} %", f"{Fu['p_ruina']:.1f} %"),
 ("Peor caso (P10)", money(H['p10']), money(Tw['p10']), money(Fu['p10'])),
 ("Mejor caso (P90)", money(H['p90']), money(Tw['p90']), money(Fu['p90'])),
]
for i,(lab,a,b,cc_) in enumerate(rows):
    s.cell(r,2,lab).font=f(10,True); s.cell(r,2).alignment=lft; s.cell(r,2).border=bd; s.cell(r,2).fill=fill(LGREY if i%2 else "FFFFFF")
    for j,v in enumerate((a,b,cc_)):
        cell=s.cell(r,3+j,v); cell.font=f(11,True); cell.alignment=ctr; cell.border=bd
        cell.fill=fill(GREEN if j==0 else (LGOLD if j==1 else ORANGE))
    s.row_dimensions[r].height=22; r+=1
r+=1

s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
c=s.cell(r,2,"Cómo leerlo"); c.font=f(12,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft; r+=1
notes=[
 f"Las tres casi nunca quiebran: tienes ventaja real, y apostar por partido (no todo junto) diversifica el riesgo entre {nbet} apuestas.",
 "½ Kelly: el más estable. Mediana ~S/278, casi imposible perder, pero el techo es más bajo. Ideal si dudas algo de tus probabilidades.",
 "⅔ Kelly: punto medio. Sube la mediana a ~S/356 y el mejor caso casi se duplica, sin meter riesgo de quiebra. Buen balance para 'moderado-agresivo'.",
 "Kelly 100%: mayor crecimiento (mediana ~S/474, P90 S/2.265) pero aparece 1,3% de riesgo de quiebra y el peor caso baja a S/59. Una pick llega a pedir el 82% de la banca — demasiado para un solo partido.",
 "Regla práctica: cuanto MENOS confíes en que tus probabilidades son exactas, MÁS baja la fracción. Si tu 85% real fuera 70%, Kelly 100% pasaría a perder dinero; ½ Kelly te protege de ese error.",
]
for t in notes:
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    c=s.cell(r,2,"•  "+t); c.font=f(10); c.alignment=lft; s.row_dimensions[r].height=30; r+=1
r+=1
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
c=s.cell(r,2,"Recomendación: ⅔ Kelly encaja con tu perfil 'entre moderado y agresivo' — más crecimiento que ½ Kelly y aún 0% de quiebra. Reserva Kelly 100% solo si estás muy seguro de tus probabilidades.")
c.font=f(10,True,"006100"); c.fill=fill(GREEN); c.alignment=lft; c.border=bd; s.row_dimensions[r].height=34

wb.move_sheet("Comparación Kelly", -(len(wb.sheetnames)-2))
wb.save("Escalera_Mundial_41.xlsx")
print("done")

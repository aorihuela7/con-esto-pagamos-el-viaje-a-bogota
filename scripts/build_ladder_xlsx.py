import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# id: (pick, partido, cuota, miP, fecha)
P = {
1:("México gana","México vs Sudáfrica",1.45,0.75,"11/06"),
2:("Suiza gana","Catar vs Suiza",1.26,0.87,"13/06"),
3:("Escocia gana","Haití vs Escocia",1.47,0.65,"17/06"),
4:("Turquía gana","Australia vs Turquía",1.80,0.78,"13/06"),
5:("Alemania gana","Alemania vs Curazao",1.05,0.98,"14/06"),
6:("España gana","España vs Cabo Verde",1.10,0.98,"15/06"),
7:("Bélgica 1X","Bélgica vs Egipto",1.19,0.77,"18/06"),
8:("Uruguay gana","Arabia Saudita vs Uruguay",1.50,0.6667,"15/06"),
9:("Francia gana","Francia vs Senegal",1.47,0.71,"16/06"),
10:("Noruega gana","Irak vs Noruega",1.23,0.87,"16/06"),
11:("Argentina gana","Argentina vs Argelia",1.42,0.68,"18/06"),
12:("Austria gana","Austria vs Jordan",1.34,0.82,"16/06"),
13:("Portugal gana","Portugal vs RD Congo",1.28,0.85,"17/06"),
14:("Ghana 1X","Ghana vs Panamá",1.29,0.80,"17/06"),
15:("Colombia gana","Uzbekistán vs Colombia",1.42,0.92,"17/06"),
16:("Rep. Checa 1X","Rep. Checa vs Sudáfrica",1.27,0.88,"18/06"),
17:("Suiza 1X","Suiza vs Bosnia",1.17,0.76,"18/06"),
18:("EE.UU. 1X","EE.UU. vs Australia",1.22,0.68,"19/06"),
19:("Marruecos gana","Escocia vs Marruecos",2.05,0.82,"19/06"),
20:("Brasil gana","Brasil vs Haití",1.06,0.99,"19/06"),
21:("Ecuador gana","Ecuador vs Curazao",1.27,0.95,"20/06"),
22:("Japón 2X","Túnez vs Japón",1.20,0.8333,"20/06"),
23:("España gana","España vs Arabia Saudita",1.13,0.885,"21/06"),
24:("Bélgica gana","Bélgica vs Irán",1.47,0.73,"21/06"),
25:("Uruguay gana","Uruguay vs Cabo Verde",1.47,0.76,"21/06"),
26:("Egipto 2X","Nueva Zelanda vs Egipto",1.19,0.87,"21/06"),
27:("Argentina gana","Argentina vs Austria",1.67,0.55,"22/06"),
28:("Francia gana","Francia vs Irak",1.12,0.96,"22/06"),
29:("Argelia 2X","Jordan vs Argelia",1.18,0.71,"25/06"),
30:("Portugal gana","Portugal vs Uzbekistán",1.27,0.95,"23/06"),
31:("Inglaterra gana","Inglaterra vs Ghana",1.32,0.82,"23/06"),
32:("Croacia gana","Panamá vs Croacia",1.50,0.84,"23/06"),
33:("Colombia gana","Colombia vs RD Congo",1.52,0.75,"23/06"),
34:("Brasil gana","Escocia vs Brasil",1.40,0.92,"24/06"),
35:("Marruecos gana","Marruecos vs Haití",1.26,0.93,"24/06"),
36:("Corea 2X","Sudáfrica vs Corea del Sur",1.27,0.80,"24/06"),
37:("C. Marfil gana","Curazao vs Costa de Marfil",1.34,0.84,"25/06"),
38:("Senegal gana","Senegal vs Irak",1.47,0.73,"26/06"),
39:("Bélgica gana","Nueva Zelanda vs Bélgica",1.30,0.86,"26/06"),
40:("Inglaterra gana","Panamá vs Inglaterra",1.20,0.95,"27/06"),
41:("Argentina gana","Jordan vs Argentina",1.20,0.91,"27/06"),
}
NOVALUE = {22, 11, 3, 27, 7, 17, 29, 18}  # 8 picks sin valor

def dkey(d): dd, mm = d.split("/"); return (int(mm), int(dd), )
order = sorted(P, key=lambda i: (dkey(P[i][4]), i))

# reference: super-combinada
prod_odds_all = 1.0; prod_p_all = 1.0
prod_odds_val = 1.0; prod_p_val = 1.0
for i in P:
    o, p = P[i][2], P[i][3]
    prod_odds_all *= o; prod_p_all *= p
    if i not in NOVALUE:
        prod_odds_val *= o; prod_p_val *= p

AR = "Arial"
NAVY="1F3864"; BLUE="2E5496"; GOLD="BF9000"
LGREY="F2F2F2"; LBLUE="DDEBF7"; LGOLD="FFF2CC"; RED="F4CCCC"; YEL="FFFDE7"; GREEN="C6EFCE"
thin=Side(style="thin",color="BFBFBF")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=Workbook()
s=wb.active; s.title="Escalera"; s.sheet_view.showGridLines=False
# cols: A sp,B #,C Fecha,D Partido/Pick,E Cuota,F Mi P,G Banca inicio,H Apostar S/,I ¿Ganó?,J Retorno,K Retirar S/,L Banca cierre
widths=[2,5,7,34,7,7,13,12,9,13,12,14]
for i,w in enumerate(widths): s.column_dimensions[get_column_letter(i+1)].width=w

s.merge_cells("B2:L2")
c=s["B2"]; c.value="ESCALERA SNOWBALL — Mundial 2026  ·  los 41 partidos, fecha a fecha"
c.font=f(17,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s.row_dimensions[2].height=30
s.merge_cells("B3:L3")
c=s["B3"]; c.value="Apuesta el día 1 y la ganancia rueda al siguiente. Tú decides cuánto re-apostar (o 0 = no apostar) y cuánto retirar a bóveda."
c.font=f(10,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=22

# Parameters
r=5
s.merge_cells(f"B{r}:D{r}"); c=s[f"B{r}"]; c.value="Banca inicial (edítala):"; c.font=f(11,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft; c.border=border
cc=s.cell(r,5,100); cc.font=f(12,True,"0000FF"); cc.fill=fill(LGOLD); cc.number_format='"S/ "#,##0'; cc.alignment=ctr; cc.border=border
s.merge_cells(f"F{r}:L{r}"); c=s[f"F{r}"]; c.value="← celda amarilla = la editas tú. Las demás amarillo claro son inputs por partido."; c.font=f(9,False,"808080"); c.alignment=lft
B0=f"$E${r}"

# Header
r=7
hdr=["#","Fecha","Partido  ·  Pick","Cuota","Mi P","Banca inicio","Apostar S/","¿Ganó? 1/0","Retorno","Retirar S/","Banca cierre"]
for j,h in enumerate(hdr):
    cc=s.cell(r,2+j,h); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(BLUE); cc.alignment=ctr; cc.border=border
s.row_dimensions[r].height=28
data0=r+1
row=data0
for k,i in enumerate(order):
    pick,part,odds,p,date=P[i]; rr=row; nv=i in NOVALUE
    gprev = B0 if k==0 else f"$L${rr-1}"
    txt=f"{part}  →  {pick}" + ("   (SIN VALOR)" if nv else "")
    s.cell(rr,2,i)
    s.cell(rr,3,date)
    s.cell(rr,4,txt)
    s.cell(rr,5,odds).number_format="0.00"
    s.cell(rr,6,p).number_format="0%"
    s.cell(rr,7,f"={gprev}").number_format='"S/ "#,##0.00'        # banca inicio
    s.cell(rr,8,None)                                             # apostar (input)
    s.cell(rr,9,None)                                             # gano (input)
    s.cell(rr,10,f'=IF($H{rr}="",0,IF($I{rr}=1,$H{rr}*$E{rr},0))').number_format='"S/ "#,##0.00'  # retorno
    s.cell(rr,11,None)                                           # retirar (input)
    s.cell(rr,12,f'=$G{rr}-IF($H{rr}="",0,$H{rr})+$J{rr}-IF($K{rr}="",0,$K{rr})').number_format='"S/ "#,##0.00'
    base = RED if nv else (LGREY if k%2 else "FFFFFF")
    for col in range(2,13):
        cell=s.cell(rr,col); cell.border=border; cell.font=f(10, nv)
        cell.alignment=lft if col==4 else ctr
        cell.fill=fill(base)
    for col in (8,9,11):  # input cols
        s.cell(rr,col).fill=fill(YEL); s.cell(rr,col).font=f(10,False,"0000FF")
    row+=1
last=row-1

# Totals
row+=0
s.merge_cells(start_row=row,start_column=2,end_row=row,end_column=6)
c=s.cell(row,2,"RESULTADO"); c.font=f(11,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=lft; c.border=border
for col in range(7,13):
    s.cell(row,col).fill=fill(NAVY); s.cell(row,col).border=border
s.cell(row,8,f'=SUM(H{data0}:H{last})').number_format='"S/ "#,##0.00'  # total apostado
s.cell(row,8).font=f(10,True,"FFFFFF")
s.cell(row,11,f'=SUM(K{data0}:K{last})').number_format='"S/ "#,##0.00' # total retirado
s.cell(row,11).font=f(10,True,"FFFFFF")
s.cell(row,12,f'=L{last}').number_format='"S/ "#,##0.00'              # banca final
s.cell(row,12).font=f(11,True,"FFFFFF")
tot=row
row+=1
s.merge_cells(start_row=row,start_column=2,end_row=row,end_column=11)
c=s.cell(row,2,"RIQUEZA TOTAL  (banca final + bóveda retirada)"); c.font=f(11,True,NAVY); c.fill=fill(GOLD if False else LGOLD); c.alignment=lft; c.border=border
cc=s.cell(row,12,f'=L{last}+SUM(K{data0}:K{last})'); cc.number_format='"S/ "#,##0.00'; cc.font=f(12,True,GOLD); cc.fill=fill(LGOLD); cc.alignment=ctr; cc.border=border
rich=row
row+=2

# Legend / how-to
for line,clr in [
 ("Fila roja = una de las 8 picks que, según el análisis de valor, NO conviene apostar (cuota peor que su probabilidad real).", RED),
 ("Celdas amarillas = inputs tuyos: Apostar S/ (0 o vacío = no apuestas ese día), ¿Ganó? (1 gana / 0 pierde), Retirar S/ (mandas ganancia a bóveda).", YEL),
 ("La Banca inicio de cada partido = Banca cierre del anterior: así rueda la ganancia sola.", LBLUE),
]:
    s.merge_cells(start_row=row,start_column=2,end_row=row,end_column=12)
    c=s.cell(row,2,"   "+line); c.font=f(9); c.fill=fill(clr); c.alignment=lft; s.row_dimensions[row].height=26; row+=1

row+=1
s.merge_cells(start_row=row,start_column=2,end_row=row,end_column=12)
c=s.cell(row,2,"La 'súper combinada' de los 41 (referencia teórica)"); c.font=f(12,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft; row+=1
refs=[
 (f"Si apostaras TODO en cada partido y ganaras los 41 (cuotas × cuotas): S/100 → S/ {100*prod_odds_all:,.0f}",
  f"...pero la probabilidad de acertar los 41 es {prod_p_all*100:.4f}%  (prácticamente imposible)"),
 (f"Quitando las 8 sin valor, los 33 buenos: S/100 → S/ {100*prod_odds_val:,.0f}",
  f"...probabilidad de acertar los 33: {prod_p_val*100:.3f}%  (sigue siendo mínima)"),
]
for a,b in refs:
    s.merge_cells(start_row=row,start_column=2,end_row=row,end_column=12); c=s.cell(row,2,"•  "+a); c.font=f(10,True); c.alignment=lft; row+=1
    s.merge_cells(start_row=row,start_column=2,end_row=row,end_column=12); c=s.cell(row,2,"    "+b); c.font=f(9,False,"808080"); c.alignment=lft; row+=1
s.merge_cells(start_row=row,start_column=2,end_row=row,end_column=12)
c=s.cell(row,2,"Por eso conviene NO re-apostar el 100% siempre: retira parte a bóveda en cada paso. Así aseguras ganancia y dejas rodar solo una porción.")
c.font=f(9,False,"808080"); c.alignment=lft; s.row_dimensions[row].height=26

s.freeze_panes="B8"
wb.save("Escalera_Mundial_41.xlsx")
print("saved", f"all41 odds prod={prod_odds_all:.0f} p={prod_p_all*100:.5f}% | val33 odds={prod_odds_val:.0f} p={prod_p_val*100:.4f}%")

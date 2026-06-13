from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

P = {
1:("México gana","México vs Sudáfrica",1.45,0.75,"11/06"),
2:("Suiza gana","Catar vs Suiza",1.26,0.87,"13/06"),
3:("Escocia gana","Haití vs Escocia",1.47,0.65,"17/06"),
4:("Turquía gana","Australia vs Turquía",1.80,0.78,"13/06"),
5:("Alemania gana","Alemania vs Curazao",1.05,0.98,"14/06"),
6:("España gana","España vs Cabo Verde",1.10,0.98,"15/06"),
7:("Bélgica 1X","Bélgica vs Egipto",1.19,0.77,"18/06"),
8:("Uruguay gana","Arabia Saudita vs Uruguay",1.50,0.6667,"15/06"),
9:("Francia gana","Francia vs Senegal",1.47,0.71,"16/06"),
10:("Noruega gana","Irak vs Noruega",1.23,0.87,"16/06"),
11:("Argentina gana","Argentina vs Argelia",1.42,0.68,"18/06"),
12:("Austria gana","Austria vs Jordan",1.34,0.82,"16/06"),
13:("Portugal gana","Portugal vs RD Congo",1.28,0.85,"17/06"),
14:("Ghana 1X","Ghana vs Panamá",1.29,0.80,"17/06"),
15:("Colombia gana","Uzbekistán vs Colombia",1.42,0.92,"17/06"),
16:("Rep. Checa 1X","Rep. Checa vs Sudáfrica",1.27,0.88,"18/06"),
17:("Suiza 1X","Suiza vs Bosnia",1.17,0.76,"18/06"),
18:("EE.UU. 1X","EE.UU. vs Australia",1.22,0.68,"19/06"),
19:("Marruecos gana","Escocia vs Marruecos",2.05,0.82,"19/06"),
20:("Brasil gana","Brasil vs Haití",1.06,0.99,"19/06"),
21:("Ecuador gana","Ecuador vs Curazao",1.27,0.95,"20/06"),
22:("Japón 2X","Túnez vs Japón",1.20,0.8333,"20/06"),
23:("España gana","España vs Arabia Saudita",1.13,0.885,"21/06"),
24:("Bélgica gana","Bélgica vs Irán",1.47,0.73,"21/06"),
25:("Uruguay gana","Uruguay vs Cabo Verde",1.47,0.76,"21/06"),
26:("Egipto 2X","Nueva Zelanda vs Egipto",1.19,0.87,"21/06"),
27:("Argentina gana","Argentina vs Austria",1.67,0.55,"22/06"),
28:("Francia gana","Francia vs Irak",1.12,0.96,"22/06"),
29:("Argelia 2X","Jordan vs Argelia",1.18,0.71,"25/06"),
30:("Portugal gana","Portugal vs Uzbekistán",1.27,0.95,"23/06"),
31:("Inglaterra gana","Inglaterra vs Ghana",1.32,0.82,"23/06"),
32:("Croacia gana","Panamá vs Croacia",1.50,0.84,"23/06"),
33:("Colombia gana","Colombia vs RD Congo",1.52,0.75,"23/06"),
34:("Brasil gana","Escocia vs Brasil",1.40,0.92,"24/06"),
35:("Marruecos gana","Marruecos vs Haití",1.26,0.93,"24/06"),
36:("Corea 2X","Sudáfrica vs Corea del Sur",1.27,0.80,"24/06"),
37:("C. Marfil gana","Curazao vs Costa de Marfil",1.34,0.84,"25/06"),
38:("Senegal gana","Senegal vs Irak",1.47,0.73,"26/06"),
39:("Bélgica gana","Nueva Zelanda vs Bélgica",1.30,0.86,"26/06"),
40:("Inglaterra gana","Panamá vs Inglaterra",1.20,0.95,"27/06"),
41:("Argentina gana","Jordan vs Argentina",1.20,0.91,"27/06"),
}
NOVALUE={22,11,3,27,7,17,29,18}

def dkey(d): dd,mm=d.split("/"); return (int(mm),int(dd))
# group value picks by date
days={}
for i in sorted(P,key=lambda i:(dkey(P[i][4]),i)):
    if i in NOVALUE: continue
    days.setdefault(P[i][4],[]).append(i)
day_order=sorted(days,key=dkey)

AR="Arial"
NAVY="1F3864"; BLUE="2E5496"; GOLD="BF9000"; TEAL="1F6E6E"
LGREY="F2F2F2"; LBLUE="DDEBF7"; LGOLD="FFF2CC"; YEL="FFFDE7"; GREEN="C6EFCE"
thin=Side(style="thin",color="BFBFBF")
med=Side(style="medium",color=NAVY)
bd=Border(left=thin,right=thin,top=thin,bottom=thin)
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=Workbook(); s=wb.active; s.title="Combinadas por día"; s.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[2,40,12,16,3]): s.column_dimensions[col].width=w

s.merge_cells("B2:D2")
c=s["B2"]; c.value="COMBINADAS POR DÍA — Mundial 2026"; c.font=f(16,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s.row_dimensions[2].height=28
s.merge_cells("B3:D3")
c=s["B3"]; c.value="Una combinada por día (solo picks con valor). Montos calculados reinvirtiendo el 100% y asumiendo que ganas. Edita el % o el ¿Ganó? cuando quieras."
c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=30

r=5
s.merge_cells(f"B{r}:C{r}"); c=s[f"B{r}"]; c.value="Banca inicial (edítala):"; c.font=f(11,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft; c.border=bd
cc=s.cell(r,4,100); cc.font=f(12,True,"0000FF"); cc.fill=fill(LGOLD); cc.number_format='"S/ "#,##0'; cc.alignment=ctr; cc.border=bd
B0=f"$D${r}"
r+=2

prev_close=B0
for di,date in enumerate(day_order,1):
    ids=days[date]; n=len(ids)
    # box header
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
    c=s.cell(r,2,f"DÍA {di}  ·  {date}   —   Combinada de {n} partido"+("s" if n>1 else ""))
    c.font=f(12,True,"FFFFFF"); c.fill=fill(TEAL); c.alignment=lft; s.row_dimensions[r].height=20
    r+=1
    # leg subheader
    for j,h in enumerate(["Partido  ·  Pick","Cuota","Mi P"]):
        cc=s.cell(r,2+j,h); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(BLUE); cc.alignment=ctr; cc.border=bd
    r+=1
    leg_first=r
    for i in ids:
        pick,part,odds,p,_=P[i]
        s.cell(r,2,f"{part}  →  {pick}").font=f(10); s.cell(r,2).alignment=lft; s.cell(r,2).border=bd
        s.cell(r,3,odds).number_format="0.00"; s.cell(r,3).alignment=ctr; s.cell(r,3).border=bd; s.cell(r,3).font=f(10)
        s.cell(r,4,p).number_format="0%"; s.cell(r,4).alignment=ctr; s.cell(r,4).border=bd; s.cell(r,4).font=f(10)
        r+=1
    leg_last=r-1
    crange=f"C{leg_first}:C{leg_last}"; prange=f"D{leg_first}:D{leg_last}"

    def srow(label, value, fmt=None, edit=False, big=False, labelfill=LGREY, valfill=None):
        global r
        s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
        c=s.cell(r,2,label); c.font=f(10,big,NAVY if big else "000000"); c.fill=fill(labelfill); c.alignment=lft; c.border=bd
        s.cell(r,3).fill=fill(labelfill); s.cell(r,3).border=bd
        cc=s.cell(r,4,value)
        if fmt: cc.number_format=fmt
        cc.alignment=ctr; cc.border=bd
        cc.font=f(11 if big else 10, True if (edit or big) else False, "0000FF" if edit else ("FFFFFF" if big else "000000"))
        cc.fill=fill(valfill if valfill else (YEL if edit else ("FFFFFF")))
        if big: cc.fill=fill(GOLD)
        cur=r; r+=1; return cur

    rc_cuota=srow("Cuota combinada", f"=PRODUCT({crange})", "0.00")
    rc_prob =srow("Prob. combinada (tus P)", f"=PRODUCT({prange})", "0.0%")
    rc_bini =srow("Banca inicio del día", f"={prev_close}", '"S/ "#,##0.00')
    # --- Kelly suggestions sub-table (% banca -> monto) ---
    base=f"((D{rc_prob}*D{rc_cuota}-1)/(D{rc_cuota}-1))"
    s.cell(r,2,"Sugerencia Kelly").font=f(10,True,NAVY); s.cell(r,2).fill=fill(LBLUE); s.cell(r,2).alignment=lft; s.cell(r,2).border=bd
    hc=s.cell(r,3,"% banca"); hc.font=f(9,True,NAVY); hc.fill=fill(LBLUE); hc.alignment=ctr; hc.border=bd
    hm=s.cell(r,4,"Monto"); hm.font=f(9,True,NAVY); hm.fill=fill(LBLUE); hm.alignment=ctr; hm.border=bd
    r+=1
    for lab,frac in [("½ Kelly (seguro)",0.5),("⅔ Kelly (medio)",2/3),("Kelly 100% (agresivo)",1.0)]:
        lc=s.cell(r,2,lab); lc.font=f(10); lc.alignment=lft; lc.border=bd; lc.fill=fill(LBLUE)
        cp=s.cell(r,3,f"=MAX(0,MIN(1,{base}*{frac}))"); cp.number_format="0.0%"; cp.alignment=ctr; cp.border=bd; cp.font=f(10); cp.fill=fill(LBLUE)
        cm=s.cell(r,4,f"=D{rc_bini}*C{r}"); cm.number_format='"S/ "#,##0.00'; cm.alignment=ctr; cm.border=bd; cm.font=f(10,True); cm.fill=fill(LBLUE)
        r+=1
    rc_pct  =srow("% a apostar  (EDITABLE)", 1.0, "0%", edit=True)
    rc_ap   =srow("Monto apostado", f"=D{rc_bini}*D{rc_pct}", '"S/ "#,##0.00')
    rc_win  =srow("¿Ganó la combinada?  (EDITABLE 1/0)", 1, "0", edit=True)
    rc_ret  =srow("Retorno", f'=IF(D{rc_win}=1,D{rc_ap}*D{rc_cuota},0)', '"S/ "#,##0.00')
    rc_close=srow("BANCA CIERRE DEL DÍA", f"=D{rc_bini}-D{rc_ap}+D{rc_ret}", '"S/ "#,##0.00', big=True)
    prev_close=f"$D${rc_close}"
    r+=1  # gap between boxes

# final summary
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
c=s.cell(r,2,"BANCA FINAL (proyección)"); c.font=f(12,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=lft; c.border=bd
s.cell(r,3).fill=fill(NAVY); s.cell(r,3).border=bd
cc=s.cell(r,4,f"={prev_close}"); cc.number_format='"S/ "#,##0.00'; cc.font=f(12,True,GOLD); cc.fill=fill(LGOLD); cc.alignment=ctr; cc.border=bd
r+=2
note=("Aviso realista: con 100% reinvertido y asumiendo que ganas TODOS los días, la banca se dispara — "
      "pero la probabilidad de encadenar los 18 días es ínfima. La columna '½ Kelly' es lo que la matemática "
      "sugiere apostar de verdad según tus probabilidades; suele ser mucho menos del 100%. Usa ½ Kelly como guía "
      "y baja el % los días de combinada arriesgada.")
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
c=s.cell(r,2,note); c.font=f(9,False,"808080"); c.alignment=lft; s.row_dimensions[r].height=70

wb.save("Combinadas_por_Dia.xlsx")
print("saved", len(day_order),"dias")

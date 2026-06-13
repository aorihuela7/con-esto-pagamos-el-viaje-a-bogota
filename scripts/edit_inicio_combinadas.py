from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from copy import copy

miP={1:0.75,2:0.87,3:0.65,4:0.78,5:0.98,6:0.98,7:0.77,8:0.6667,9:0.71,10:0.87,
11:0.68,12:0.82,13:0.85,14:0.80,15:0.92,16:0.88,17:0.76,18:0.68,19:0.82,20:0.99,
21:0.95,22:0.8333,23:0.885,24:0.73,25:0.76,26:0.87,27:0.55,28:0.96,29:0.71,30:0.95,
31:0.82,32:0.84,33:0.75,34:0.92,35:0.93,36:0.80,37:0.84,38:0.73,39:0.86,40:0.95,41:0.91}
YEL="FFFDE7"
def fill(c): return PatternFill("solid",fgColor=c)

wb=load_workbook("Sistema_Mundial_2026_COMPLETO.xlsx")
s=wb["Inicio · Combinadas"]
s.column_dimensions["E"].width=14
s.column_dimensions["F"].width=12

first_leg=None
for r in range(1,s.max_row+1):
    a=s.cell(r,1).value
    if isinstance(a,str) and a.startswith("COMBINADA"):
        for mr in list(s.merged_cells.ranges):
            if mr.min_row==r and mr.max_row==r and mr.min_col==1 and mr.max_col==4:
                s.unmerge_cells(str(mr)); s.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); break
        continue
    if a=="#":
        hd=s.cell(r,4)
        for col,txt in ((5,"Prob. de la cuota"),(6,"Mi prob.")):
            c=s.cell(r,col,txt); c.font=copy(hd.font); c.fill=copy(hd.fill); c.alignment=copy(hd.alignment); c.border=copy(hd.border)
        first_leg=None
        continue
    if isinstance(a,int):
        if first_leg is None: first_leg=r
        last_leg=r
        ds=s.cell(r,4)
        ce=s.cell(r,5,f"=1/D{r}"); ce.number_format="0.0%"
        ce.border=copy(ds.border); ce.alignment=copy(ds.alignment); ce.font=copy(ds.font); ce.fill=copy(ds.fill)
        cf=s.cell(r,6,miP.get(a,"")); cf.number_format="0.0%"
        cf.border=copy(ds.border); cf.alignment=copy(ds.alignment)
        cf.font=Font(name="Arial",size=ds.font.sz,color="0000FF"); cf.fill=fill(YEL)
        continue
    if s.cell(r,3).value=="Cuota total" and first_leg:
        dt=s.cell(r,4)
        ce=s.cell(r,5,f"=PRODUCT(E{first_leg}:E{last_leg})"); ce.number_format="0.000%"
        cf=s.cell(r,6,f"=PRODUCT(F{first_leg}:F{last_leg})"); cf.number_format="0.000%"
        for c in (ce,cf):
            c.font=copy(dt.font); c.alignment=copy(dt.alignment); c.border=copy(dt.border); c.fill=copy(dt.fill)
        continue

wb.save("Sistema_Mundial_2026_COMPLETO.xlsx")
print("columnas E/F agregadas a Inicio · Combinadas")

import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d = json.load(open("cmp.json"))
od, fk, dates = d["od"], d["fk"], d["dates"]
A, B = d["A"], d["B"]
b100, bk = d["b100det"], d["bkdet"]

AR="Arial"; NAVY="1F3864"; BLUE="2E5496"; TEAL="1F6E6E"; GOLD="BF9000"
LGREY="F2F2F2"; LBLUE="DDEBF7"; LGOLD="FFF2CC"; GREEN="C6EFCE"; RED="F4CCCC"
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=load_workbook("Combinadas_por_Dia.xlsx")
if "Comparación" in wb.sheetnames: del wb["Comparación"]
s=wb.create_sheet("Comparación")
s.sheet_view.showGridLines=False
for col,w in zip("ABCDEFG",[2,7,9,12,12,18,18]): s.column_dimensions[col].width=w

s.merge_cells("B2:G2")
c=s["B2"]; c.value="½ KELLY  vs  APOSTAR 100%"; c.font=f(16,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s.row_dimensions[2].height=28
s.merge_cells("B3:G3")
c=s["B3"]; c.value="Las dos estrategias arrancan con S/100. Abajo: primero el cuento de hadas (ganas todo), luego la realidad (200.000 simulaciones con TUS probabilidades)."
c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=28

# ---- Escenario 1: gano todo, dia a dia ----
r=5
s.merge_cells(f"B{r}:G{r}")
c=s[f"B{r}"]; c.value="Escenario A — \"Gano TODOS los días\" (la fantasía)"; c.font=f(12,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft
r+=1
hdr=["Día","Fecha","Cuota comb","½Kelly %","Banca 100%","Banca ½ Kelly"]
for j,h in enumerate(hdr):
    cc=s.cell(r,2+j,h); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(BLUE); cc.alignment=ctr; cc.border=bd
s.row_dimensions[r].height=26
r+=1
b1=100.0; b2=100.0
for i in range(len(od)):
    b1*=od[i]; b2*=(1+fk[i]*(od[i]-1))
    vals=[i+1, dates[i], round(od[i],3), fk[i], b1, b2]
    for j,v in enumerate(vals):
        cc=s.cell(r,2+j,v); cc.border=bd; cc.font=f(10); cc.alignment=ctr
        if j==3: cc.number_format="0.0%"
        if j==2: cc.number_format="0.00"
        if j==4: cc.number_format='"S/ "#,##0'
        if j==5: cc.number_format='"S/ "#,##0'
        cc.fill=fill(LGREY if i%2 else "FFFFFF")
    s.cell(r,6).fill=fill(LGOLD); s.cell(r,7).fill=fill(GREEN)
    r+=1
# total row
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
c=s.cell(r,2,"Si ganaras los 16 días seguidos →"); c.font=f(10,True,NAVY); c.fill=fill(LGOLD); c.alignment=lft; c.border=bd
for col in (3,4,5): s.cell(r,col).fill=fill(LGOLD); s.cell(r,col).border=bd
s.cell(r,6,round(b100)).number_format='"S/ "#,##0'; s.cell(r,6).font=f(11,True,GOLD); s.cell(r,6).fill=fill(LGOLD); s.cell(r,6).alignment=ctr; s.cell(r,6).border=bd
s.cell(r,7,round(bk)).number_format='"S/ "#,##0'; s.cell(r,7).font=f(11,True,"006100"); s.cell(r,7).fill=fill(GREEN); s.cell(r,7).alignment=ctr; s.cell(r,7).border=bd
r+=1
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
c=s.cell(r,2,"Aquí 100% gana más... PERO la probabilidad de encadenar los 16 días es solo 0,44%. Mira lo que pasa en la realidad ↓")
c.font=f(9,False,"808080"); c.alignment=lft; s.row_dimensions[r].height=18
r+=2

# ---- Escenario 2: realista ----
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
c=s.cell(r,2,"Escenario B — Realidad: 200.000 mundiales simulados con TUS probabilidades"); c.font=f(12,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft
r+=1
hdr2=["Qué mide","Apostar 100%","½ Kelly","Qué significa"]
spans=[(2,2),(3,3),(4,4),(5,7)]
for (a,b),h in zip(spans,hdr2):
    s.merge_cells(start_row=r,start_column=a,end_row=r,end_column=b)
    cc=s.cell(r,a,h); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(BLUE); cc.alignment=ctr; cc.border=bd
    for col in range(a,b+1): s.cell(r,col).border=bd
s.row_dimensions[r].height=24
r+=1
rows=[
 ("Banca final típica (mediana)", f"S/ {A['med']:.0f}", f"S/ {B['med']:.0f}", "Lo que te pasa la MITAD de las veces", True),
 ("Prob. de terminar ganando", f"{A['p_prof']:.1f} %", f"{B['p_prof']:.1f} %", "Acabar con más de tus S/100", True),
 ("Prob. de quebrar (perderlo casi todo)", f"{A['p_ruina']:.1f} %", f"{B['p_ruina']:.0f} %", "Quedarte casi en cero", False),
 ("Peor caso (P10)", f"S/ {A['p10']:.0f}", f"S/ {B['p10']:.0f}", "1 de cada 10 mundiales sale así o peor", True),
 ("Mejor caso (P90)", f"S/ {A['p90']:.0f}", f"S/ {B['p90']:.0f}", "1 de cada 10 sale así o mejor", False),
 ("Promedio", f"S/ {A['mean']:.0f}", f"S/ {B['mean']:.0f}", "El de 100% sube solo por el jackpot rarísimo", False),
]
for i,(lab,va,vb,note,good) in enumerate(rows):
    s.cell(r,2,lab).font=f(10,True); s.cell(r,2).alignment=lft; s.cell(r,2).border=bd; s.cell(r,2).fill=fill(LGREY if i%2 else "FFFFFF")
    ca=s.cell(r,3,va); ca.font=f(10,True); ca.alignment=ctr; ca.border=bd; ca.fill=fill(RED)
    cb=s.cell(r,4,vb); cb.font=f(10,True); cb.alignment=ctr; cb.border=bd; cb.fill=fill(GREEN)
    s.merge_cells(start_row=r,start_column=5,end_row=r,end_column=7)
    cn=s.cell(r,5,note); cn.font=f(9,False,"808080"); cn.alignment=lft; cn.border=bd
    for col in (6,7): s.cell(r,col).border=bd; s.cell(r,col).fill=fill(LGREY if i%2 else "FFFFFF")
    s.row_dimensions[r].height=20; r+=1
r+=1

# takeaway
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
c=s.cell(r,2,"La conclusión"); c.font=f(12,True,NAVY); c.fill=fill(LGOLD); c.alignment=lft; r+=1
takes=[
 "Apostar 100% cada día: el 99,6% de las veces terminas en CERO. Solo el 0,4% pega el premio gordo. Es lotería, no estrategia.",
 "½ Kelly: el 88% de las veces terminas GANANDO, nunca quiebras, y en un mundial típico duplicas (mediana S/219).",
 "Kelly no busca el premio máximo; busca que SIGAS en el juego y crezcas parejo. Por eso gana en la realidad aunque pierda en la fantasía.",
 "Tu archivo ya trae la columna ½ Kelly por día: úsala como el monto a apostar, y baja el % si dudas de tus probabilidades.",
]
for t in takes:
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
    c=s.cell(r,2,"•  "+t); c.font=f(10); c.alignment=lft; s.row_dimensions[r].height=22; r+=1

# move sheet right after the daily sheet
wb.move_sheet("Comparación", -(len(wb.sheetnames)-2))
wb.save("Combinadas_por_Dia.xlsx")
print("done")

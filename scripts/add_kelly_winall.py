import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

d=json.load(open("cmp.json"))
od=d["od"]; pd=d["pd"]; dates=d["dates"]
N=len(od)

AR="Arial"; NAVY="1F3864"; BLUE="2E5496"; GREEN="1F6E6E"; GOLD="BF9000"
LGREY="F2F2F2"; LBLUE="DDEBF7"; LGOLD="FFF2CC"; GREENF="C6EFCE"
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=load_workbook("Combinadas_por_Dia.xlsx")

variants=[("½ Kelly — gana todo",0.5,GREENF,"006100"),
          ("⅔ Kelly — gana todo",2/3,LGOLD,"7F6000"),
          ("Kelly 100% — gana todo",1.0,"FCE4D6","843C0C")]

for title,frac,accent,acol in variants:
    if title in wb.sheetnames: del wb[title]
    s=wb.create_sheet(title); s.sheet_view.showGridLines=False
    for col,w in zip("ABCDEFGHIJ",[2,6,8,11,10,9,15,14,14,16]): s.column_dimensions[col].width=w
    s.merge_cells("B2:J2")
    c=s["B2"]; c.value=f"{title.upper()}  ·  combinada por día, asumiendo que GANAS cada día"
    c.font=f(14,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s.row_dimensions[2].height=26
    s.merge_cells("B3:J3")
    c=s["B3"]; c.value="Cada día apuestas SOLO la fracción de Kelly de la banca (no el 100%). El resto queda guardado. Si ganas, sumas la ganancia y rueda al día siguiente."
    c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=26
    # banca inicial
    r=5
    s.merge_cells(f"B{r}:D{r}"); c=s[f"B{r}"]; c.value="Banca inicial:"; c.font=f(11,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft; c.border=bd
    cc=s.cell(r,5,100); cc.font=f(12,True,"0000FF"); cc.fill=fill(LGOLD); cc.number_format='"S/ "#,##0'; cc.alignment=ctr; cc.border=bd
    B0=f"$E${r}"
    r=7
    hdr=["Día","Fecha","Cuota comb","Prob comb","Kelly %","Banca inicio","Apostado","Ganancia","Banca cierre"]
    for j,h in enumerate(hdr):
        cell=s.cell(r,2+j,h); cell.font=f(9,True,"FFFFFF"); cell.fill=fill(BLUE); cell.alignment=ctr; cell.border=bd
    s.row_dimensions[r].height=28
    r+=1; first=r
    for i in range(N):
        rr=r
        gprev=B0 if i==0 else f"$J${rr-1}"
        s.cell(rr,2,i+1)
        s.cell(rr,3,dates[i])
        s.cell(rr,4,round(od[i],3)).number_format="0.000"
        s.cell(rr,5,pd[i]).number_format="0.0%"
        s.cell(rr,6,f"=MAX(0,MIN(1,((E{rr}*D{rr}-1)/(D{rr}-1))*{frac}))").number_format="0.0%"
        s.cell(rr,7,f"={gprev}").number_format='"S/ "#,##0.00'
        s.cell(rr,8,f"=G{rr}*F{rr}").number_format='"S/ "#,##0.00'
        s.cell(rr,9,f"=H{rr}*(D{rr}-1)").number_format='"S/ "#,##0.00'
        s.cell(rr,10,f"=G{rr}+I{rr}").number_format='"S/ "#,##0.00'
        for col in range(2,11):
            cell=s.cell(rr,col); cell.border=bd; cell.font=f(10); cell.alignment=ctr
            cell.fill=fill(LGREY if i%2 else "FFFFFF")
        s.cell(rr,10).fill=fill(accent); s.cell(rr,10).font=f(10,True,acol)
        r+=1
    last=r-1
    # totals
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
    c=s.cell(r,2,"BANCA FINAL (si ganas los 16 días)"); c.font=f(11,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=lft; c.border=bd
    for col in range(8,11): s.cell(r,col).fill=fill(NAVY); s.cell(r,col).border=bd
    s.cell(r,8,f"=SUM(H{first}:H{last})").number_format='"S/ "#,##0.00'; s.cell(r,8).font=f(10,True,"FFFFFF")
    s.cell(r,9,f"=SUM(I{first}:I{last})").number_format='"S/ "#,##0.00'; s.cell(r,9).font=f(10,True,"FFFFFF")
    s.cell(r,10,f"=J{last}").number_format='"S/ "#,##0.00'; s.cell(r,10).font=f(12,True,GOLD); s.cell(r,10).fill=fill(LGOLD)
    r+=2
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=10)
    c=s.cell(r,2,"Nota: 'Apostado' es solo la fracción Kelly; el resto de la banca no se arriesga ese día. Por eso crece más lento que apostar el 100%, pero te protege cuando un día falla (este cuadro asume que ganas todos).")
    c.font=f(9,False,"808080"); c.alignment=lft; s.row_dimensions[r].height=30
    s.freeze_panes="B8"

wb.save("Combinadas_por_Dia.xlsx")
print("done")

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d=json.load(open("system_data.json"))
picks=d["picks"]; pm=d["pm"]; combos=d["combos"]; excl=d["excl"]; mc=d["mc"]
byid={p[0]:p for p in picks}
F="Arial"
WHITE="FFFFFF"; NAVY="1F4E78"; BLUE="0000FF"; GREEN="006100"
hdr=Font(name=F,bold=True,color=WHITE,size=11)
hfill=PatternFill("solid",start_color=NAVY)
sub=Font(name=F,bold=True,size=11)
base=Font(name=F,size=10)
inp=Font(name=F,size=10,color=BLUE,bold=True)
inpfill=PatternFill("solid",start_color="FFF2CC")
center=Alignment(horizontal="center",vertical="center")
left=Alignment(horizontal="left",vertical="center")
right=Alignment(horizontal="right",vertical="center")
thin=Side(style="thin",color="D0D0D0")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
band_fill={"A":"E2EFDA","M":"FFF2CC","L":"FCE4D6"}

wb=Workbook()

# ===== Probabilidades (editable) =====
sp=wb.create_sheet("Probabilidades")
heads=["#","Pick","Partido","Cuota","P mercado","MI PROB (editable)","Mi cuota justa","En # combos"]
for j,h in enumerate(heads,1):
    c=sp.cell(1,j,h); c.font=hdr; c.fill=hfill; c.alignment=center; c.border=border
prow={}
for i,p in enumerate(picks,2):
    pid=p[0]; prow[pid]=i
    sp.cell(i,1,pid).alignment=center
    sp.cell(i,2,p[1])
    sp.cell(i,3,p[2])
    sp.cell(i,4,p[3]).number_format="0.00"
    sp.cell(i,5,f"=1/D{i}").number_format="0.0%"
    mc_cell=sp.cell(i,6,round(pm[str(pid)],4)); mc_cell.number_format="0.0%"; mc_cell.font=inp; mc_cell.fill=inpfill
    sp.cell(i,7,f"=1/F{i}").number_format="0.00"
    sp.cell(i,8,f"{combos and ''}{len(combos)-excl[str(pid)]}/{len(combos)}").alignment=center
    for j in range(1,9):
        cc=sp.cell(i,j); cc.border=border
        if cc.font.color is None or j!=6: cc.font=base if j!=6 else inp
        if j in(1,4,5,7,8): cc.alignment=center
sp.freeze_panes="A2"
sp.cell(len(picks)+3,2,"Edita SOLO la columna amarilla 'MI PROB' con tu lectura de cada partido. Todo lo demás se recalcula solo.").font=Font(name=F,italic=True,size=10,color=BLUE)
for col,w in zip("ABCDEFGH",[5,26,30,8,11,18,13,12]): sp.column_dimensions[col].width=w

# ===== Sistema =====
ss=wb.create_sheet("Sistema")
ss.cell(1,1,"SISTEMA DE COMBINADAS — barra (anclas cortas + medias + lotería)").font=Font(name=F,bold=True,size=14)
ss.cell(2,1,"Apuesta (azul) editable. EV>0 = combinada con valor según TU probabilidad.").font=Font(name=F,italic=True,size=10)
row=4
ev_cells=[]; stake_cells=[]
for ci in combos:
    band=ci["band"]; legs=ci["legs"]
    c=ss.cell(row,1,f"COMBINADA {ci['name']}  ·  {ci['n']} patas"); c.font=Font(name=F,bold=True,size=12,color=WHITE)
    for col in range(1,5): ss.cell(row,col).fill=hfill
    ss.merge_cells(start_row=row,start_column=1,end_row=row,end_column=4); row+=1
    for hi,h in enumerate(["Pick","Partido","Cuota","Mi P"],1):
        cc=ss.cell(row,hi,h); cc.font=sub; cc.border=border; cc.alignment=center
    row+=1; first=row
    for pid in legs:
        p=byid[pid]; r=prow[pid]
        ss.cell(row,1,p[1])
        ss.cell(row,2,p[2])
        oc=ss.cell(row,3,f"='Probabilidades'!D{r}"); oc.number_format="0.00"; oc.alignment=center
        pc=ss.cell(row,4,f"='Probabilidades'!F{r}"); pc.number_format="0.0%"; pc.alignment=center
        for col in range(1,5): ss.cell(row,col).font=base; ss.cell(row,col).border=border
        ss.cell(row,1).fill=PatternFill("solid",start_color=band_fill[band])
        row+=1
    last=row-1
    cuota_c=f"C{row}"; ss.cell(row,1,"Cuota total").font=sub; ss.cell(row,1).alignment=right
    cc=ss.cell(row,3,f"=ROUND(PRODUCT(C{first}:C{last}),2)"); cc.font=Font(name=F,bold=True); cc.number_format="0.00"; cc.alignment=center; row+=1
    prob_c=f"C{row}"; ss.cell(row,1,"P(gana) según tu prob").font=sub; ss.cell(row,1).alignment=right
    cc=ss.cell(row,3,f"=PRODUCT(D{first}:D{last})"); cc.font=Font(name=F,bold=True); cc.number_format="0.00%"; cc.alignment=center; row+=1
    stake_c=f"C{row}"; ss.cell(row,1,"Apuesta (S/) — editable").font=sub; ss.cell(row,1).alignment=right
    cc=ss.cell(row,3,ci["stake"]); cc.font=inp; cc.fill=inpfill; cc.number_format="0"; cc.alignment=center; stake_cells.append(stake_c); row+=1
    pago_c=f"C{row}"; ss.cell(row,1,"Pago si gana (S/)").font=sub; ss.cell(row,1).alignment=right
    cc=ss.cell(row,3,f"={cuota_c}*{stake_c}"); cc.font=Font(name=F,bold=True,color=GREEN); cc.number_format='"S/ "#,##0.00'; cc.alignment=center; row+=1
    ev_c=f"C{row}"; ss.cell(row,1,"EV (S/)").font=sub; ss.cell(row,1).alignment=right
    cc=ss.cell(row,3,f"={prob_c}*{pago_c}-{stake_c}"); cc.font=Font(name=F,bold=True); cc.number_format='"S/ "#,##0.00;[Red]"S/ -"#,##0.00'; cc.alignment=center; ev_cells.append(ev_c); row+=1
    ss.cell(row,1,"¿+EV?").font=sub; ss.cell(row,1).alignment=right
    cc=ss.cell(row,3,f'=IF({ev_c}>0,"SÍ ✓","no")'); cc.font=Font(name=F,bold=True); cc.alignment=center; row+=2
for col,w in zip("ABCD",[24,30,12,10]): ss.column_dimensions[col].width=w

# ===== Resumen =====
sr=wb.create_sheet("Resumen",0)
sr.cell(1,1,"Sistema de combinadas — Mundial 2026").font=Font(name=F,bold=True,size=16)
sr.cell(2,1,"Base: cuotas de mercado · 14 combinadas (5 anclas, 5 medias, 4 lotería)").font=Font(name=F,italic=True,size=11)
r=4
sr.cell(r,1,"TOTALES (en vivo, según tus ediciones)").font=sub; r+=1
sum_stake="+".join(f"Sistema!{s}" for s in stake_cells)
sum_ev="+".join(f"Sistema!{e}" for e in ev_cells)
rows=[("Apuesta total (S/)",f"={sum_stake}",'"S/ "#,##0'),
      ("EV total del portafolio (S/)",f"={sum_ev}",'"S/ "#,##0.00;[Red]"S/ -"#,##0.00'),
      ("# combinadas con +EV",f'=COUNTIF(Sistema!{ev_cells[0]}:{ev_cells[0]},">0")',"0")]
# COUNTIF over scattered cells not trivial; build SUMPRODUCT of >0 flags
flags="+".join(f"IF(Sistema!{e}>0,1,0)" for e in ev_cells)
rows[2]=("# combinadas con +EV (de 14)",f"={flags}","0")
for lab,fml,fmt in rows:
    sr.cell(r,1,lab).font=base
    cc=sr.cell(r,2,fml); cc.font=Font(name=F,bold=True); cc.number_format=fmt; cc.alignment=center; r+=1
r+=1
sr.cell(r,1,"BASELINE con probabilidades de mercado (simulación 200k, sin tu ventaja)").font=sub; r+=1
base_rows=[("P(que alguna combinada gane)",f"{mc['any_win_pct']}%"),
           ("P(la ronda sea rentable, neto>0)",f"{mc['p_profit']}%"),
           ("Neto promedio por ronda",f"S/ {mc['avg_net']}"),
           ("Apuesta total (provisional)",f"S/ {mc['total_stake']:.0f}")]
for lab,val in base_rows:
    sr.cell(r,1,lab).font=base; sr.cell(r,2,val).font=base; sr.cell(r,2).alignment=center; r+=1
r+=1
for n in ["CÓMO USARLO:",
          "1) Ve a 'Probabilidades' y ajusta la columna amarilla MI PROB con tu lectura de cada partido.",
          "2) Mira 'Sistema': las combinadas con EV>0 (¿+EV? = SÍ) son donde tu lectura le gana al mercado.",
          "3) Ajusta las apuestas (azul) en 'Sistema'. Aquí arriba ves el EV total en vivo.",
          "4) En vivo usa 'Cashout y Hedge' para decidir cuándo cerrar o asegurar.",
          "NOTA: el baseline de arriba es estático (cuotas de mercado). El EV en vivo usa TU prob."]:
    cc=sr.cell(r,1,n); cc.font=Font(name=F,bold=n.endswith(":"),size=10); sr.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
sr.column_dimensions["A"].width=46; sr.column_dimensions["B"].width=18

# ===== Cashout y Hedge =====
ch=wb.create_sheet("Cashout y Hedge")
ch.cell(1,1,"Gestión en vivo: Cashout y Cobertura (Hedge)").font=Font(name=F,bold=True,size=14)
r=3
ch.cell(r,1,"REGLA DE CASHOUT (heurística)").font=sub; r+=1
for n in ["El cashout convierte una combinada 'viva' en valor asegurado. Idea: compara lo ofrecido contra lo que esperas si sigues.",
          "Valor esperado de seguir = (prob. conjunta de las patas restantes) × (pago potencial).",
          "Cierra (cashout) cuando: el cashout ofrecido ≥ ese valor esperado, O cuando ya asegura tu objetivo de ganancia y no quieres arriesgar.",
          "Con aversión al riesgo (proteger banca), conviene cerrar ANTES de lo que diría el EV puro.",
          "Regla práctica simple: si te quedan ≤2 patas y el cashout supera tu umbral de ganancia, cierra."]:
    cc=ch.cell(r,1,"• "+n); cc.font=base; ch.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); r+=1
r+=1
ch.cell(r,1,"CALCULADORA DE HEDGE (última pata viva)").font=sub; r+=1
ch.cell(r,1,"Cuando una combinada llega viva a su última pata, puedes asegurar ganancia apostando al resultado contrario.").font=base
ch.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7); r+=2
calc=[("Pago potencial si gana la combinada (W, S/)",1000,'"S/ "#,##0'),
      ("Cuota del resultado CONTRARIO (o)",2.50,"0.00"),
      ("Lo ya apostado en la combinada (S, S/)",20,'"S/ "#,##0')]
addr={}
for lab,val,fmt in calc:
    ch.cell(r,1,lab).font=base
    cc=ch.cell(r,2,val); cc.font=inp; cc.fill=inpfill; cc.number_format=fmt; cc.alignment=center
    addr[lab]=f"B{r}"; r+=1
W=addr[calc[0][0]]; O=addr[calc[1][0]]; S=addr[calc[2][0]]
res=[("Cuánto apostar al contrario (hedge = W/o)",f"={W}/{O}",'"S/ "#,##0.00'),
     ("Ganancia GARANTIZADA pase lo que pase",f"={W}*(1-1/{O})-{S}",'"S/ "#,##0.00;[Red]"S/ -"#,##0.00')]
for lab,fml,fmt in res:
    ch.cell(r,1,lab).font=sub
    cc=ch.cell(r,2,fml); cc.font=Font(name=F,bold=True,color=GREEN); cc.number_format=fmt; cc.alignment=center; r+=1
ch.column_dimensions["A"].width=46; ch.column_dimensions["B"].width=16

if "Sheet" in wb.sheetnames: del wb["Sheet"]
wb.save("Sistema_Combinadas_Mundial.xlsx")
print("saved")

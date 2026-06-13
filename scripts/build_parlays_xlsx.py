import json, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d=json.load(open("parlays.json"))
data=d["data"]; P={int(k):v for k,v in d["P"].items()}
N=d["N"]; STAKE=d["stake"]; cover=d["cover"]; minhit=d["minhit"]
cost=d["cost"]; p_any=d["p_any"]; e_gross=d["e_gross"]; e_net=d["e_net"]; p_win=d["p_win"]; sims=d["sims"]; e_hits=d["e_hits"]

AR="Arial"; NAVY="1F3864"; BLUE="2E5496"; TEAL="1F6E6E"; GOLD="BF9000"
LGREY="F2F2F2"; LBLUE="DDEBF7"; LGOLD="FFF2CC"; GREEN="C6EFCE"; RED="F4CCCC"; ORANGE="FCE4D6"
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=Workbook()

# ============ SHEET 1: RESUMEN Y RIESGO ============
s=wb.active; s.title="Resumen y riesgo"; s.sheet_view.showGridLines=False
for col,w in zip("ABCDEFGH",[2,7,11,11,12,11,11,3]): s.column_dimensions[col].width=w
s.merge_cells("B2:G2")
c=s["B2"]; c.value="18 COMBINADAS PARA GANAR +S/5.000  ·  S/30 cada una"; c.font=f(15,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s.row_dimensions[2].height=28
s.merge_cells("B3:G3")
c=s["B3"]; c.value=("Cada combinada paga más de S/5.000 de ganancia con solo S/30. Están mezcladas para que si falla 1 o 2 partidos NUNCA se caigan todas, "
 "y aún con 3 fallos casi siempre sobrevive alguna. Inversión total: S/"+str(cost)+" (18 × S/30).")
c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=32

# stake editable
r=5
s.merge_cells(f"B{r}:C{r}"); c=s[f"B{r}"]; c.value="Apuesta por combinada:"; c.font=f(10,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft; c.border=bd
cc=s.cell(r,4,STAKE); cc.font=f(11,True,"0000FF"); cc.fill=fill(LGOLD); cc.number_format='"S/ "#,##0'; cc.alignment=ctr; cc.border=bd
ST=f"$D${r}"

r=7
hdr=["#","Cuota comb.","Paga (S/)","Ganancia (S/)","Prob. mía","Partidos"]
for j,h in enumerate(hdr):
    cc=s.cell(r,2+j,h); cc.font=f(9,True,"FFFFFF"); cc.fill=fill(BLUE); cc.alignment=ctr; cc.border=bd
s.row_dimensions[r].height=28
r+=1; first=r
for k,dd in enumerate(data,1):
    s.cell(r,2,k).font=f(10,True)
    s.cell(r,3,round(dd["odds"],1)).number_format="0.0"
    s.cell(r,4,f"={ST}*C{r}").number_format='"S/ "#,##0'
    s.cell(r,5,f"={ST}*C{r}-{ST}").number_format='"S/ "#,##0'
    s.cell(r,6,dd["prob"]).number_format="0.0%"
    s.cell(r,7,dd["n"])
    for col in range(2,8):
        cell=s.cell(r,col); cell.border=bd; cell.alignment=ctr
        if not cell.font.bold: cell.font=f(10)
        cell.fill=fill(LGREY if k%2 else "FFFFFF")
    s.cell(r,5).font=f(10,True,"006100")
    r+=1
last=r-1
# total invested
s.merge_cells(f"B{r}:C{r}"); c=s.cell(r,2,"Inversión total (18 combinadas)"); c.font=f(10,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=lft; c.border=bd
cc=s.cell(r,4,f"={ST}*{N}"); cc.number_format='"S/ "#,##0'; cc.font=f(10,True,"FFFFFF"); cc.fill=fill(NAVY); cc.border=bd
for col in (3,5,6,7): s.cell(r,col).fill=fill(NAVY); s.cell(r,col).border=bd
r+=2

# resiliencia
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
c=s.cell(r,2,"Qué tan blindadas están (si fallan partidos)"); c.font=f(12,True,NAVY); c.fill=fill(LGOLD); c.alignment=lft; r+=1
res=[
 (f"Si falla 1 partido", "GARANTIZADO: siempre sobreviven varias combinadas. Ningún partido está en todas.", GREEN),
 (f"Si fallan 2 partidos", "GARANTIZADO: siempre sobrevive al menos una combinada (probado con las 496 parejas posibles).", GREEN),
 (f"Si fallan 3 partidos", f"En el {100*cover['3'][0]/cover['3'][1]:.0f}% de los casos sobrevive al menos una. Solo combinaciones muy puntuales de 3 fallos las tumban todas.", LGOLD),
 (f"Para tumbar las 18", f"Se necesitan al menos {minhit} partidos fallando a la vez, y justo los que se cruzan. Un par de sorpresas no te deja en cero.", LBLUE),
]
for lab,txt,col in res:
    s.merge_cells(start_row=r,start_column=3,end_row=r,end_column=7)
    cl=s.cell(r,2,lab); cl.font=f(10,True); cl.alignment=lft; cl.border=bd; cl.fill=fill(col)
    ct=s.cell(r,3,txt); ct.font=f(9); ct.alignment=lft; ct.border=bd; ct.fill=fill(col)
    s.row_dimensions[r].height=30; r+=1
r+=1

# monte carlo realista
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
c=s.cell(r,2,"La verdad realista (200.000+ mundiales simulados con TUS probabilidades)"); c.font=f(12,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft; r+=1
mc=[
 ("Costo total", f"S/ {cost}", "Lo que pones en las 18 combinadas."),
 ("Prob. de que AL MENOS una acierte", f"{p_any:.0f}%", "Las otras veces (las 18 fallan) pierdes los S/"+str(cost)+"."),
 ("Pago bruto esperado", f"S/ {e_gross:,.0f}", "Promedio a largo plazo; en la práctica es 0 o un golpe de +S/5.000."),
 ("Resultado neto esperado", f"+S/ {e_net:,.0f}", "Positivo SOLO si tus probabilidades son reales. Si son optimistas, baja."),
]
for lab,val,note in mc:
    cl=s.cell(r,2,lab); cl.font=f(10,True); cl.alignment=lft; cl.border=bd; cl.fill=fill(LGREY)
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
    cv=s.cell(r,4,val); cv.font=f(11,True,"006100"); cv.alignment=ctr; cv.border=bd
    s.merge_cells(start_row=r,start_column=5,end_row=r,end_column=7)
    cn=s.cell(r,5,note); cn.font=f(9,False,"808080"); cn.alignment=lft; cn.border=bd
    s.row_dimensions[r].height=26; r+=1
r+=1
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
c=s.cell(r,2,"Sé honesto contigo: son combinadas de 14-17 partidos, loterías de premio gordo. Lo más probable (≈65%) es que un mundial cualquiera las tumbe todas y pierdas los S/"+str(cost)+". El diseño NO sube tu chance de ganar el premio: reparte el riesgo para que una mala tarde no borre TODO, y para que cada acierto pague +S/5.000. Si buscas crecer la banca de forma estable, las hojas de Escalera/Kelly son mejores; esto es el billete de lotería con ventaja.")
c.font=f(9,False,"808080"); c.alignment=lft; s.row_dimensions[r].height=70
s.freeze_panes="B8"

# ============ SHEET 2: DETALLE (legs por combinada) ============
s2=wb.create_sheet("Detalle por combinada"); s2.sheet_view.showGridLines=False
for col,w in zip("ABCDE",[2,9,30,22,9]): s2.column_dimensions[col].width=w
s2.merge_cells("B2:E2")
c=s2["B2"]; c.value="DETALLE DE CADA COMBINADA  ·  los partidos que la forman"; c.font=f(14,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s2.row_dimensions[2].height=26
r=4
for k,dd in enumerate(data,1):
    legs=sorted(dd["legs"], key=lambda i:(P[i][4][3:],P[i][4][:2]))  # by fecha dd/mm
    s2.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    gan=STAKE*dd["odds"]-STAKE
    c=s2.cell(r,2,f"COMBINADA #{k}   ·   cuota {dd['odds']:.0f}   ·   paga S/{STAKE*dd['odds']:,.0f}   ·   ganancia S/{gan:,.0f}   ·   {dd['n']} partidos   ·   prob mía {dd['prob']*100:.1f}%")
    c.font=f(10,True,"FFFFFF"); c.fill=fill(TEAL); c.alignment=lft; r+=1
    for j,h in enumerate(["Fecha","Partido","Apuesta","Cuota"]):
        cc=s2.cell(r,2+j,h); cc.font=f(8,True,"FFFFFF"); cc.fill=fill(BLUE); cc.alignment=ctr; cc.border=bd
    r+=1
    for i in legs:
        lab,part,od,mp,fe=P[i]
        s2.cell(r,2,fe).alignment=ctr; s2.cell(r,2).font=f(9)
        s2.cell(r,3,part).alignment=lft; s2.cell(r,3).font=f(9)
        s2.cell(r,4,lab).alignment=lft; s2.cell(r,4).font=f(9,False,"1F3864")
        s2.cell(r,5,od).alignment=ctr; s2.cell(r,5).font=f(9); s2.cell(r,5).number_format="0.00"
        for col in range(2,6): s2.cell(r,col).border=bd; s2.cell(r,col).fill=fill("FFFFFF")
        r+=1
    r+=1

# ============ SHEET 3: MAPA DE DIVERSIFICACION ============
s3=wb.create_sheet("Mapa de diversificacion"); s3.sheet_view.showGridLines=False
used=sorted({i for dd in data for i in dd["legs"]}, key=lambda i:(P[i][4][3:],P[i][4][:2]))
s3.column_dimensions["A"].width=2
s3.column_dimensions["B"].width=8; s3.column_dimensions["C"].width=26; s3.column_dimensions["D"].width=20
for k in range(N): s3.column_dimensions[get_column_letter(5+k)].width=4.2
s3.merge_cells(start_row=2,start_column=2,end_row=2,end_column=4+N)
c=s3.cell(2,2,"MAPA DE DIVERSIFICACIÓN  ·  ✓ = el partido está en esa combinada"); c.font=f(13,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s3.row_dimensions[2].height=24
s3.merge_cells(start_row=3,start_column=2,end_row=3,end_column=4+N)
c=s3.cell(3,2,"Mira cualquier fila: ningún partido tiene ✓ en todas las columnas. Por eso un fallo nunca tumba las 18."); c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s3.row_dimensions[3].height=22
r=5
s3.cell(r,2,"Fecha").font=f(8,True,"FFFFFF"); s3.cell(r,2).fill=fill(BLUE)
s3.cell(r,3,"Partido").font=f(8,True,"FFFFFF"); s3.cell(r,3).fill=fill(BLUE)
s3.cell(r,4,"Apuesta").font=f(8,True,"FFFFFF"); s3.cell(r,4).fill=fill(BLUE)
for col in (2,3,4): s3.cell(r,col).alignment=ctr; s3.cell(r,col).border=bd
for k in range(N):
    cc=s3.cell(r,5+k,k+1); cc.font=f(8,True,"FFFFFF"); cc.fill=fill(BLUE); cc.alignment=ctr; cc.border=bd
s3.row_dimensions[r].height=18
r+=1
legsets=[set(dd["legs"]) for dd in data]
for i in used:
    lab,part,od,mp,fe=P[i]
    s3.cell(r,2,fe).font=f(8); s3.cell(r,2).alignment=ctr
    s3.cell(r,3,part).font=f(8); s3.cell(r,3).alignment=lft
    s3.cell(r,4,lab).font=f(8,False,"1F3864"); s3.cell(r,4).alignment=lft
    cnt=0
    for k in range(N):
        if i in legsets[k]:
            cc=s3.cell(r,5+k,"✓"); cc.font=f(9,True,"006100"); cc.fill=fill(GREEN); cnt+=1
        else:
            cc=s3.cell(r,5+k,""); cc.fill=fill("FFFFFF")
        cc.alignment=ctr; cc.border=bd
    for col in (2,3,4): s3.cell(r,col).border=bd; s3.cell(r,col).fill=fill(LGREY)
    r+=1
# bottom: count per parlay
s3.cell(r,4,"# partidos →").font=f(8,True); s3.cell(r,4).alignment=Alignment(horizontal="right")
for k in range(N):
    cc=s3.cell(r,5+k,data[k]["n"]); cc.font=f(8,True,"FFFFFF"); cc.fill=fill(NAVY); cc.alignment=ctr; cc.border=bd
s3.freeze_panes="E6"

wb.save("Combinadas_5000.xlsx")
print("saved", N,"combinadas")

import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

d=json.load(open("cmp3.json")); R=d["res"]; nd=d["ndays"]; fmax=d["fmax"]
H,Tw,Fu=R["half"],R["twothirds"],R["full"]

AR="Arial"; NAVY="1F3864"; BLUE="2E5496"
LGREY="F2F2F2"; LBLUE="DDEBF7"; LGOLD="FFF2CC"; GREEN="C6EFCE"; ORANGE="FCE4D6"
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)

wb=load_workbook("Combinadas_por_Dia.xlsx")
name="Comparación Kelly (por día)"
if name in wb.sheetnames: del wb[name]
s=wb.create_sheet(name); s.sheet_view.showGridLines=False
for col,w in zip("ABCDEF",[2,34,15,15,15,3]): s.column_dimensions[col].width=w

s.merge_cells("B2:E2")
c=s["B2"]; c.value="¿QUÉ FRACCIÓN DE KELLY?  —  Combinada por día"; c.font=f(15,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; s.row_dimensions[2].height=28
s.merge_cells("B3:E3")
c=s["B3"]; c.value=(f"Apuestas una combinada cada día ({nd} días), arriesgando una fracción de Kelly de la banca en cada una. "
 "200.000 mundiales simulados con TUS probabilidades, banca inicial S/100.")
c.font=f(9,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; s.row_dimensions[3].height=32

r=5
hdr=["Métrica","½ Kelly (conservador)","⅔ Kelly (medio)","Kelly 100% (agresivo)"]
fl=[BLUE,GREEN,LGOLD,ORANGE]
for j,h in enumerate(hdr):
    cc=s.cell(r,2+j,h); cc.font=f(10,True,"FFFFFF" if j==0 else "000000"); cc.fill=fill(fl[j]); cc.alignment=ctr; cc.border=bd
s.row_dimensions[r].height=30
r+=1
def m(x): return f"S/ {x:.0f}"
rows=[
 ("Banca típica (mediana)", m(H['med']), m(Tw['med']), m(Fu['med'])),
 ("Prob. de terminar ganando", f"{H['p_prof']:.1f} %", f"{Tw['p_prof']:.1f} %", f"{Fu['p_prof']:.1f} %"),
 ("Prob. de quebrar (perder 90%+)", f"{H['p_ruina']:.1f} %", f"{Tw['p_ruina']:.1f} %", f"{Fu['p_ruina']:.1f} %"),
 ("Peor caso (P10)", m(H['p10']), m(Tw['p10']), m(Fu['p10'])),
 ("Mejor caso (P90)", m(H['p90']), m(Tw['p90']), m(Fu['p90'])),
]
for i,(lab,a,b,cc_) in enumerate(rows):
    s.cell(r,2,lab).font=f(10,True); s.cell(r,2).alignment=lft; s.cell(r,2).border=bd; s.cell(r,2).fill=fill(LGREY if i%2 else "FFFFFF")
    for j,v in enumerate((a,b,cc_)):
        cell=s.cell(r,3+j,v); cell.font=f(11,True); cell.alignment=ctr; cell.border=bd
        cell.fill=fill(GREEN if j==0 else (LGOLD if j==1 else ORANGE))
    s.row_dimensions[r].height=22; r+=1
r+=1

s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
c=s.cell(r,2,"Cómo leerlo"); c.font=f(12,True,NAVY); c.fill=fill(LBLUE); c.alignment=lft; r+=1
notes=[
 "Combinada por día = más riesgo por apuesta que ir partido a partido (multiplicas varias probabilidades), por eso las medianas son algo más bajas y la varianza más alta.",
 "½ Kelly: el más estable. Mediana ~S/219, 88% de ganar, 0% de quiebra. El techo más bajo (P90 S/462).",
 "⅔ Kelly: sube la mediana a ~S/260 y el mejor caso a ~S/706, sin riesgo de quiebra. Buen balance moderado-agresivo.",
 "Kelly 100%: mayor crecimiento (mediana ~S/314, P90 S/1.420) pero asoma 1,1% de quiebra y el peor caso cae a S/48.",
 "Comparado con la Escalera partido a partido, aquí TODO rinde un poco menos: agrupar partidos en una combinada concentra el riesgo. Si quieres máxima estabilidad, la escalera gana; si quieres menos apuestas y más adrenalina, la combinada por día.",
]
for t in notes:
    s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    c=s.cell(r,2,"•  "+t); c.font=f(10); c.alignment=lft; s.row_dimensions[r].height=30; r+=1
r+=1
s.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
c=s.cell(r,2,"Recomendación: ⅔ Kelly para tu perfil 'entre moderado y agresivo'. La columna ½ Kelly de la hoja por día ya te da el monto; multiplícalo por 1,33 si quieres ⅔ Kelly.")
c.font=f(10,True,"006100"); c.fill=fill(GREEN); c.alignment=lft; c.border=bd; s.row_dimensions[r].height=34

wb.move_sheet(name, -(len(wb.sheetnames)-2))
wb.save("Combinadas_por_Dia.xlsx")
print("done")

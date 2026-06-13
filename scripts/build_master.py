from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

AR="Arial"; NAVY="1F3864"; BLUE="2E5496"; TEAL="1F6E6E"; GOLD="BF9000"
LBLUE="DDEBF7"; LGOLD="FFF2CC"; GREEN="C6EFCE"
def f(sz=10,b=False,color="000000"): return Font(name=AR,size=sz,bold=b,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
lft=Alignment(horizontal="left",vertical="center",wrap_text=True)
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)

def copy_sheet(src,dst):
    dst.sheet_view.showGridLines=src.sheet_view.showGridLines
    for k,dim in src.column_dimensions.items():
        dst.column_dimensions[k].width=dim.width
        dst.column_dimensions[k].hidden=dim.hidden
    for k,dim in src.row_dimensions.items():
        dst.row_dimensions[k].height=dim.height
    for row in src.iter_rows():
        for c in row:
            nc=dst.cell(row=c.row,column=c.column,value=c.value)
            if c.has_style:
                nc.font=copy(c.font); nc.fill=copy(c.fill); nc.border=copy(c.border)
                nc.alignment=copy(c.alignment); nc.number_format=c.number_format
    for mr in list(src.merged_cells.ranges):
        dst.merge_cells(str(mr))
    if src.freeze_panes: dst.freeze_panes=src.freeze_panes

# (file, source sheet, target name)
plan=[
 ("Combinadas_Mundial_2026.xlsx","Resumen",     "Inicio · Resumen"),
 ("Combinadas_Mundial_2026.xlsx","Combinadas",  "Inicio · Combinadas"),
 ("Combinadas_Mundial_2026.xlsx","Cobertura",   "Inicio · Cobertura"),
 ("Combinadas_por_Dia.xlsx","Combinadas por día",        "Kelly · Por día"),
 ("Combinadas_por_Dia.xlsx","Comparación Kelly (por día)","Kelly · Comparación"),
 ("Combinadas_por_Dia.xlsx","½ Kelly — gana todo",       "Kelly ½ · gana todo"),
 ("Combinadas_por_Dia.xlsx","⅔ Kelly — gana todo",       "Kelly ⅔ · gana todo"),
 ("Combinadas_por_Dia.xlsx","Kelly 100% — gana todo",    "Kelly 100% · gana todo"),
 ("Combinadas_5000.xlsx","Resumen y riesgo",        "5000 · Resumen y riesgo"),
 ("Combinadas_5000.xlsx","Detalle por combinada",   "5000 · Detalle"),
 ("Combinadas_5000.xlsx","Mapa de diversificacion", "5000 · Mapa"),
]

wb=Workbook(); idx=wb.active; idx.title="Portada"; idx.sheet_view.showGridLines=False
for col,w in zip("ABCD",[2,30,60,2]): idx.column_dimensions[col].width=w
idx.merge_cells("B2:C2")
c=idx["B2"]; c.value="SISTEMA DE APUESTAS · MUNDIAL 2026"; c.font=f(16,True,"FFFFFF"); c.fill=fill(NAVY); c.alignment=ctr; idx.row_dimensions[2].height=32
idx.merge_cells("B3:C3")
c=idx["B3"]; c.value="Todo el trabajo reunido en un solo archivo. Tres bloques, una pestaña por hoja."; c.font=f(10,False,"FFFFFF"); c.fill=fill(BLUE); c.alignment=ctr; idx.row_dimensions[3].height=22
r=5
blocks=[
 ("1 · LAS COMBINADAS INICIALES", TEAL, [
   ("Inicio · Resumen","Visión general del primer sistema de combinadas."),
   ("Inicio · Combinadas","Las combinadas armadas al principio, con su cobertura."),
   ("Inicio · Cobertura","Qué partidos cubre cada combinada."),
 ]),
 ("2 · KELLY (lo último que trabajamos)", GOLD, [
   ("Kelly · Por día","Combinada por día con la sugerencia de Kelly ½, ⅔ y 100%."),
   ("Kelly · Comparación","½ vs ⅔ vs 100% Kelly: mediana, prob. de ganar, de quebrar, P10, P90."),
   ("Kelly ½ · gana todo","Trayectoria de la banca si apuestas solo ½ Kelly y ganas siempre."),
   ("Kelly ⅔ · gana todo","Igual pero con ⅔ Kelly."),
   ("Kelly 100% · gana todo","Igual pero apostando el Kelly completo."),
 ]),
 ("3 · COMBINADAS DE +S/5.000", NAVY, [
   ("5000 · Resumen y riesgo","18 combinadas de S/30 que pagan +S/5.000, con su análisis de riesgo."),
   ("5000 · Detalle","Los partidos que forman cada una de las 18 combinadas."),
   ("5000 · Mapa","Mapa visual de diversificación (✓ por combinada)."),
 ]),
]
for title,col,items in blocks:
    idx.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
    c=idx.cell(r,2,title); c.font=f(12,True,"FFFFFF"); c.fill=fill(col); c.alignment=lft; idx.row_dimensions[r].height=24; r+=1
    for name,desc in items:
        cl=idx.cell(r,2,name); cl.font=f(10,True,NAVY); cl.alignment=lft; cl.border=bd; cl.fill=fill(LBLUE)
        cd=idx.cell(r,3,desc); cd.font=f(9); cd.alignment=lft; cd.border=bd
        idx.row_dimensions[r].height=24; r+=1
    r+=1
idx.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
c=idx.cell(r,2,"Recordatorio: las hojas de Kelly hacen crecer la banca de forma estable; las de +S/5.000 son billetes de lotería con ventaja (lo más probable es perder la apuesta, pero cada acierto paga gordo). Apuesta solo lo que puedas perder.")
c.font=f(9,False,"808080"); c.alignment=lft; idx.row_dimensions[r].height=46

for fn,src_name,tgt in plan:
    swb=load_workbook(fn)
    if src_name not in swb.sheetnames:
        print("FALTA",fn,src_name); continue
    ws=wb.create_sheet(tgt[:31])
    copy_sheet(swb[src_name],ws)
    print("ok",tgt)

wb.save("Sistema_Mundial_2026_COMPLETO.xlsx")
print("saved",len(wb.sheetnames),"hojas")
